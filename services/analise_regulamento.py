from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

from models.inscricao import normalizar_chave


CAMPO_ALIASES = {
    "inscricao_id": ["Inscrição (Código)", "Inscricao (Codigo)", "ID", "CNPJ"],
    "empresa_nome": ["Nome Fantasia", "Empresa", "Razão Social", "Razao Social"],
    "razao_social": ["Razão Social", "Razao Social"],
    "cnpj": ["CNPJ"],
    "data_submissao": ["Data de Inscricao", "Criado Em", "Created On"],
    "status_financeiro": ["Status Financeiro"],
    "classificacao_porte_maturidade": [
        "Classificacao Porte e Maturidade",
        "Porte e Maturidade",
        "Classificação Porte e Maturidade",
    ],
    "perfil_publico_alvo_declarado": [
        "Publico-Alvo",
        "Publico alvo",
        "Sua empresa e micro, pequena ou media empresa",
    ],
    "setor_classificacao": ["Setor - Classificacao", "Setor - Classificação"],
    "website_empresa": ["Website", "Site", "Rede Social", "(C016)"],
    "website_internacional_resposta": [
        "Lingua estrangeira",
        "Língua estrangeira",
        "Comprador estrangeiro",
        "(C017)",
    ],
    "website_internacional_link": ["Website", "Lingua estrangeira", "(C018)"],
    "cartao_cnpj_link": ["Cartão CNPJ", "Cartao CNPJ", "(C008)", "(C024)"],
    "participa_peiex": ["PEIEX", "(C012)"],
    "participa_consultoria_sebrae_exportacao": ["SEBRAE", "Sebrae", "(C015)"],
    "ncm_descricao_produtos": ["NCM", "(C010)"],
    "catalogo_digital_link": ["Catalogo", "Catálogo", "Flyer", "Material promocional", "(C020)"],
    "liderada_por_mulher": ["Mulher", "Liderada por mulher", "(C021)"],
    "liderada_por_pessoa_negra": ["Pessoa negra", "Preta", "Parda", "(C022)"],
    "diversidade_regional": ["Norte", "Nordeste", "DF", "Distrito Federal", "(C023)"],
    "possui_certificado_origem": ["Certificado de origem", "(C025)"],
    "certificado_origem_link": ["Certificado de origem", "(C026)"],
    "possui_certificacao_internacional": ["Certificacao internacional", "Certificação internacional", "(C027)"],
    "certificacao_internacional_link": ["Certificacao internacional", "Certificação internacional", "(C014)"],
}

CRITERIOS_BIBLIOTECA = [
    ("elegibilidade", "publico_alvo", ["publico-alvo", "publico alvo", "mpe", "micro", "pequena", "media empresa"]),
    ("elegibilidade", "pendencia_financeira", ["pendencia financeira", "inadimpl", "adimplente"]),
    ("verificacao", "cnpj", ["cnpj", "cartao cnpj", "comprovante de inscricao"]),
    ("pontuacao", "website_rede_social", ["website", "site", "rede social"]),
    ("pontuacao", "idioma_estrangeiro", ["ingles", "espanhol", "comprador estrangeiro"]),
    ("pontuacao", "peiex", ["peiex"]),
    ("pontuacao", "sebrae", ["sebrae"]),
    ("pontuacao", "ncm", ["ncm"]),
    ("pontuacao", "catalogo_material", ["catalogo", "flyer", "material promocional"]),
    ("pontuacao", "certificado_origem", ["certificado de origem"]),
    ("pontuacao", "certificacao_internacional", ["certificacao internacional", "certificado atualizado"]),
    ("pontuacao", "diversidade_genero", ["liderada por mulher", "mulher"]),
    ("pontuacao", "diversidade_racial", ["pessoa negra", "preta", "parda"]),
    ("pontuacao", "diversidade_regional", ["norte", "nordeste", "df"]),
    ("demanda", "compradores", ["compradores", "demanda", "mercado-alvo", "mercado alvo"]),
]


@dataclass
class ResumoAnaliseRegulamento:
    modo: str
    regulamento_path: str
    entrada_path: str
    saida_path: str
    sheet_utilizada: str
    total_colunas_identificadas: int
    total_mapeamentos_sugeridos: int
    total_criterios_detectados: int


def executar_analise_regulamento(
    entrada_path: str | Path,
    regulamento_path: str | Path,
    saida_path: str | Path,
    *,
    aba: str | None = None,
) -> ResumoAnaliseRegulamento:
    entrada = Path(entrada_path)
    regulamento = Path(regulamento_path)
    saida = Path(saida_path)

    texto_regulamento = extrair_texto_regulamento(regulamento)
    inspecao = inspecionar_planilha(entrada, sheet_name=aba)
    criterios = detectar_criterios(texto_regulamento)
    mapeamentos = sugerir_mapeamentos(inspecao["headers"])
    config_sugerida = montar_config_sugerida(
        sheet_name=inspecao["sheet_name"],
        mapeamentos=mapeamentos,
        criterios=criterios,
        nome_programa=regulamento.stem,
    )

    exportar_analise_regulamento(
        saida,
        texto_regulamento=texto_regulamento,
        inspecao=inspecao,
        criterios=criterios,
        mapeamentos=mapeamentos,
        config_sugerida=config_sugerida,
    )

    return ResumoAnaliseRegulamento(
        modo="analise_regulamento",
        regulamento_path=str(regulamento),
        entrada_path=str(entrada),
        saida_path=str(saida),
        sheet_utilizada=inspecao["sheet_name"],
        total_colunas_identificadas=len(inspecao["headers"]),
        total_mapeamentos_sugeridos=sum(1 for item in mapeamentos if item["header_origem"]),
        total_criterios_detectados=len(criterios),
    )


def extrair_texto_regulamento(caminho: str | Path) -> str:
    path = Path(caminho)
    extensao = path.suffix.lower()
    if extensao == ".pdf":
        reader = PdfReader(str(path))
        paginas = [pagina.extract_text() or "" for pagina in reader.pages]
        return "\n\n".join(paginas).strip()
    if extensao in {".txt", ".md", ".json"}:
        return path.read_text(encoding="utf-8")
    raise ValueError("Formato de regulamento nao suportado para analise. Use PDF, TXT, MD ou JSON.")


def inspecionar_planilha(caminho: str | Path, *, sheet_name: str | None = None) -> dict[str, Any]:
    path = Path(caminho)
    extensao = path.suffix.lower()
    if extensao == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as arquivo:
            leitor = csv.reader(arquivo)
            headers = next(leitor, [])
            amostras = []
            for indice, linha in enumerate(leitor):
                if indice >= 3:
                    break
                amostras.append(linha)
        return {"sheet_name": "csv", "headers": headers, "samples": amostras}

    if extensao not in {".xlsx", ".xlsm"}:
        raise ValueError("Formato de planilha nao suportado para analise. Use CSV ou XLSX.")

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    worksheet = _resolver_worksheet(workbook, sheet_name)
    linhas = worksheet.iter_rows(values_only=True)
    headers = list(_obter_primeira_linha_preenchida(linhas) or [])
    amostras = []
    for linha in linhas:
        if not linha or not any(celula is not None and str(celula).strip() for celula in linha):
            continue
        amostras.append(list(linha))
        if len(amostras) >= 3:
            break
    workbook.close()
    return {"sheet_name": worksheet.title, "headers": headers, "samples": amostras}


def detectar_criterios(texto_regulamento: str) -> list[dict[str, Any]]:
    texto = texto_regulamento or ""
    texto_normalizado = normalizar_chave(texto).replace("_", " ")
    resultados: list[dict[str, Any]] = []

    for categoria, criterio_id, palavras in CRITERIOS_BIBLIOTECA:
        hits = [palavra for palavra in palavras if normalizar_chave(palavra).replace("_", " ") in texto_normalizado]
        if not hits:
            continue
        trecho = _extrair_trecho(texto, hits[0])
        resultados.append(
            {
                "categoria": categoria,
                "criterio_id": criterio_id,
                "palavras_encontradas": ", ".join(hits),
                "trecho": trecho,
            }
        )
    return resultados


def sugerir_mapeamentos(headers: list[Any]) -> list[dict[str, Any]]:
    headers_texto = ["" if header is None else str(header) for header in headers]
    resultados = []

    for campo_destino, aliases in CAMPO_ALIASES.items():
        header_origem = ""
        confianca = "nao_encontrado"
        for header in headers_texto:
            header_norm = normalizar_chave(header)
            for alias in aliases:
                alias_norm = normalizar_chave(alias)
                if header_norm == alias_norm:
                    header_origem = header
                    confianca = "alta"
                    break
                if alias_norm and alias_norm in header_norm:
                    header_origem = header
                    confianca = "media"
                    break
            if header_origem:
                break

        resultados.append(
            {
                "campo_destino": campo_destino,
                "header_origem": header_origem,
                "confianca": confianca,
            }
        )

    return resultados


def montar_config_sugerida(
    *,
    sheet_name: str,
    mapeamentos: list[dict[str, Any]],
    criterios: list[dict[str, Any]],
    nome_programa: str,
) -> dict[str, Any]:
    field_map = {
        item["campo_destino"]: [item["header_origem"]]
        for item in mapeamentos
        if item["header_origem"]
    }
    return {
        "nome_programa": nome_programa,
        "versao": "rascunho",
        "origem_dados": {
            "sheet_name": sheet_name,
            "mapeamento_campos": field_map,
        },
        "criterios_detectados": criterios,
        "elegibilidade": [],
        "pontuacao": {"nota_minima_classificacao": 0, "criterios": []},
        "desempate": [],
    }


def exportar_analise_regulamento(
    caminho_saida: str | Path,
    *,
    texto_regulamento: str,
    inspecao: dict[str, Any],
    criterios: list[dict[str, Any]],
    mapeamentos: list[dict[str, Any]],
    config_sugerida: dict[str, Any],
) -> None:
    path = Path(caminho_saida)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    aba_resumo = workbook.active
    aba_resumo.title = "resumo_analise"
    _escrever_tabela(
        aba_resumo,
        ["indicador", "valor"],
        [
            {"indicador": "aba_utilizada", "valor": inspecao["sheet_name"]},
            {"indicador": "total_colunas", "valor": len(inspecao["headers"])},
            {"indicador": "total_criterios_detectados", "valor": len(criterios)},
            {
                "indicador": "total_mapeamentos_sugeridos",
                "valor": sum(1 for item in mapeamentos if item["header_origem"]),
            },
        ],
    )

    _escrever_tabela(
        workbook.create_sheet("colunas_planilha"),
        ["coluna_original", "coluna_normalizada", "amostra_1", "amostra_2", "amostra_3"],
        _linhas_colunas_planilha(inspecao["headers"], inspecao["samples"]),
    )
    _escrever_tabela(
        workbook.create_sheet("criterios_detectados"),
        ["categoria", "criterio_id", "palavras_encontradas", "trecho"],
        criterios,
    )
    _escrever_tabela(
        workbook.create_sheet("mapeamento_sugerido"),
        ["campo_destino", "header_origem", "confianca"],
        mapeamentos,
    )
    _escrever_tabela(
        workbook.create_sheet("regulamento_extraido"),
        ["linha", "texto"],
        [{"linha": indice + 1, "texto": linha} for indice, linha in enumerate((texto_regulamento or "").splitlines())],
    )
    _escrever_tabela(
        workbook.create_sheet("config_sugerida"),
        ["linha", "json"],
        [
            {"linha": indice + 1, "json": linha}
            for indice, linha in enumerate(
                json.dumps(config_sugerida, ensure_ascii=False, indent=2).splitlines()
            )
        ],
    )

    workbook.save(path)
    workbook.close()


def _linhas_colunas_planilha(headers: list[Any], samples: list[list[Any]]) -> list[dict[str, Any]]:
    linhas = []
    for indice, header in enumerate(headers):
        linha = {
            "coluna_original": header,
            "coluna_normalizada": normalizar_chave(header or f"coluna_{indice + 1}"),
            "amostra_1": samples[0][indice] if len(samples) > 0 and indice < len(samples[0]) else None,
            "amostra_2": samples[1][indice] if len(samples) > 1 and indice < len(samples[1]) else None,
            "amostra_3": samples[2][indice] if len(samples) > 2 and indice < len(samples[2]) else None,
        }
        linhas.append(linha)
    return linhas


def _escrever_tabela(worksheet, cabecalho: list[str], linhas: list[dict[str, Any]]) -> None:
    worksheet.append(cabecalho)
    for celula in worksheet[1]:
        celula.font = Font(bold=True)
    worksheet.freeze_panes = "A2"

    for linha in linhas:
        worksheet.append([linha.get(coluna) for coluna in cabecalho])

    for indice_coluna, coluna in enumerate(worksheet.columns, start=1):
        valores = ["" if celula.value is None else str(celula.value) for celula in coluna]
        largura = min(max(len(valor) for valor in valores) + 2, 80) if valores else 12
        worksheet.column_dimensions[get_column_letter(indice_coluna)].width = largura


def _resolver_worksheet(workbook, sheet_name: str | None):
    if not sheet_name:
        return workbook.active
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    desejado = normalizar_chave(sheet_name)
    for nome_aba in workbook.sheetnames:
        if normalizar_chave(nome_aba) == desejado:
            return workbook[nome_aba]
    return workbook.active


def _obter_primeira_linha_preenchida(linhas) -> tuple[Any, ...] | None:
    for linha in linhas:
        if linha is not None and any(celula is not None and str(celula).strip() for celula in linha):
            return linha
    return None


def _extrair_trecho(texto: str, termo: str, *, contexto: int = 110) -> str:
    if not texto:
        return ""

    padrao = re.escape(termo)
    match = re.search(padrao, texto, flags=re.IGNORECASE)
    if not match:
        return ""

    inicio = max(0, match.start() - contexto)
    fim = min(len(texto), match.end() + contexto)
    return " ".join(texto[inicio:fim].split())

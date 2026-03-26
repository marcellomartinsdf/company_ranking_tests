from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from models import Inscricao
from models.inscricao import normalizar_chave


def carregar_inscricoes(
    caminho: str | Path,
    *,
    sheet_name: str | None = None,
    field_map: dict[str, Any] | None = None,
) -> list[Inscricao]:
    path = Path(caminho)
    extensao = path.suffix.lower()

    if extensao == ".csv":
        return _carregar_csv(path, field_map=field_map)
    if extensao in {".xlsx", ".xlsm"}:
        return _carregar_xlsx(path, sheet_name=sheet_name, field_map=field_map)

    raise ValueError("Formato de entrada nao suportado. Use CSV ou XLSX.")


def _carregar_csv(
    path: Path,
    *,
    field_map: dict[str, Any] | None = None,
) -> list[Inscricao]:
    with path.open("r", encoding="utf-8-sig", newline="") as arquivo:
        amostra = arquivo.read(2048)
        arquivo.seek(0)
        try:
            dialect = csv.Sniffer().sniff(amostra, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel

        leitor = csv.DictReader(arquivo, dialect=dialect)
        return [
            Inscricao.from_row(_aplicar_mapeamento_campos(linha, field_map))
            for linha in leitor
            if any(linha.values())
        ]


def _carregar_xlsx(
    path: Path,
    *,
    sheet_name: str | None = None,
    field_map: dict[str, Any] | None = None,
) -> list[Inscricao]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    if _parece_workbook_saida_sistema(workbook):
        workbook.close()
        raise ValueError(
            "A planilha enviada parece ser uma saida do proprio sistema de ranqueamento. "
            "Envie a exportacao bruta do Dynamics para calcular as notas a partir do formulario."
        )
    worksheet = _resolver_worksheet(workbook, sheet_name)
    linhas = worksheet.iter_rows(values_only=True)
    cabecalho = _obter_primeira_linha_preenchida(linhas)
    if cabecalho is None:
        workbook.close()
        return []

    inscricoes: list[Inscricao] = []
    for linha in linhas:
        if linha is None or not any(celula is not None and str(celula).strip() for celula in linha):
            continue
        row_dict = {
            str(cabecalho[indice] or f"coluna_{indice + 1}"): linha[indice]
            for indice in range(len(cabecalho))
        }
        inscricoes.append(Inscricao.from_row(_aplicar_mapeamento_campos(row_dict, field_map)))

    workbook.close()
    return inscricoes


def _resolver_worksheet(workbook, sheet_name: str | None):
    if not sheet_name:
        return workbook.active

    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    nome_desejado = normalizar_chave(sheet_name)
    for nome_aba in workbook.sheetnames:
        if normalizar_chave(nome_aba) == nome_desejado:
            return workbook[nome_aba]

    return workbook.active


def _obter_primeira_linha_preenchida(linhas) -> tuple[Any, ...] | None:
    for linha in linhas:
        if linha is not None and any(celula is not None and str(celula).strip() for celula in linha):
            return linha
    return None


def _aplicar_mapeamento_campos(
    row_dict: dict[str, Any],
    field_map: dict[str, Any] | None,
) -> dict[str, Any]:
    if not field_map:
        return row_dict

    row_mapeado = dict(row_dict)
    lookup_normalizado = {
        normalizar_chave(chave): valor
        for chave, valor in row_dict.items()
        if chave is not None
    }

    for campo_destino, aliases in field_map.items():
        aliases_lista = aliases if isinstance(aliases, list) else [aliases]
        for alias in aliases_lista:
            valor = _buscar_valor_por_alias(lookup_normalizado, alias)
            if valor is None or str(valor).strip() == "":
                continue
            row_mapeado[campo_destino] = valor
            break

    return row_mapeado


def _buscar_valor_por_alias(lookup_normalizado: dict[str, Any], alias: Any) -> Any:
    chave_alias = normalizar_chave(alias)
    if not chave_alias:
        return None

    if chave_alias in lookup_normalizado:
        return lookup_normalizado[chave_alias]

    if len(chave_alias) < 12:
        return None

    tokens_alias = [token for token in chave_alias.split("_") if token]
    alias_pontuacao = any(token in {"pontuacao", "pontos"} for token in tokens_alias)
    candidatos: list[tuple[int, int, Any]] = []
    for chave_lookup, valor in lookup_normalizado.items():
        if valor is None or str(valor).strip() == "":
            continue

        if chave_alias in chave_lookup or chave_lookup in chave_alias:
            candidatos.append((len(tokens_alias), abs(len(chave_lookup) - len(chave_alias)), valor))
            continue

        if alias_pontuacao:
            continue

        tokens_lookup = [token for token in chave_lookup.split("_") if token]
        tokens_comuns = len(set(tokens_alias).intersection(tokens_lookup))
        if tokens_comuns >= 3 and tokens_comuns / max(len(tokens_alias), 1) >= 0.6:
            candidatos.append((tokens_comuns, abs(len(chave_lookup) - len(chave_alias)), valor))

    if not candidatos:
        return None

    candidatos.sort(key=lambda item: (-item[0], item[1]))
    return candidatos[0][2]


def _parece_workbook_saida_sistema(workbook) -> bool:
    nomes_normalizados = {normalizar_chave(nome_aba) for nome_aba in workbook.sheetnames}
    abas_saida = {
        "ranking_final",
        "avaliacao_por_criterio",
        "pendencias_revisao",
        "inscricoes_brutas",
    }
    return abas_saida.issubset(nomes_normalizados)

from __future__ import annotations

from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import Inscricao, ResultadoInscricao


COR_AZUL = "2046F2"
COR_AZUL_ESCURO = "132B8F"
COR_AZUL_CLARO = "EEF3FF"
COR_AZUL_CLARO_2 = "F8FAFF"
COR_VERDE = "DDF6DE"
COR_VERDE_TEXTO = "0B6B22"
COR_AMARELO = "FFF3C8"
COR_AMARELO_TEXTO = "8A5B00"
COR_VERMELHO = "FCE1DC"
COR_VERMELHO_TEXTO = "A4372D"
COR_CINZA = "E8EDF8"
COR_CINZA_TEXTO = "5B678D"
COR_BORDA = "D7DFF5"

COLUNAS_LARGAS = {
    "justificativa",
    "motivo_status_final",
    "resumo_classificacao",
    "valor_observado",
    "criterio_nome",
    "empresa_nome",
    "razao_social",
}
COLUNAS_STATUS = {"status_final", "resultado", "status_final_inscricao"}
COLUNAS_BOOLEANAS = {
    "elegivel",
    "elegibilidade_pendente_revisao",
    "nota_minima_atendida",
    "demanda_atendida",
    "demanda_pendente_revisao",
    "revisao_humana_pendente",
    "contabilizar_na_nota",
}
COLUNAS_NUMERICAS_DECIMAIS = {"pontuacao", "pontuacao_maxima", "pontuacao_total", "pontuacao_sugerida"}
COLUNAS_NUMERICAS_INTEIRAS = {"ordem", "classificacao"}
COLUNAS_DATA = {"data_submissao"}


class ExportadorExcel:
    def __init__(self, regulamento: dict[str, Any] | None = None) -> None:
        self.regulamento = regulamento or {}
        self.config_saida = dict(self.regulamento.get("saida_excel") or {})

    def exportar(
        self,
        caminho_saida: str | Path,
        inscricoes: list[Inscricao],
        ranking: list[ResultadoInscricao],
    ) -> None:
        path = Path(caminho_saida)
        path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        planilha_ranking = workbook.active
        planilha_ranking.title = "ranking_final"
        self._configurar_aparencia_planilha(planilha_ranking)
        self._preencher_ranking_final(planilha_ranking, ranking)

        if not self._somente_ranking_final():
            planilha_avaliacao = workbook.create_sheet("avaliacao_por_criterio")
            self._configurar_aparencia_planilha(planilha_avaliacao)
            self._preencher_avaliacao_por_criterio(planilha_avaliacao, ranking)

            planilha_pendencias = workbook.create_sheet("pendencias_revisao")
            self._configurar_aparencia_planilha(planilha_pendencias)
            self._preencher_pendencias_revisao(planilha_pendencias, ranking)

            planilha_bruta = workbook.create_sheet("inscricoes_brutas")
            self._configurar_aparencia_planilha(planilha_bruta)
            self._preencher_inscricoes_brutas(planilha_bruta, inscricoes)
            planilha_bruta.sheet_state = "hidden"

        self._aplicar_cores_abas(workbook)
        workbook.active = 0

        workbook.save(path)
        workbook.close()

    def _somente_ranking_final(self) -> bool:
        return bool(self.config_saida.get("somente_ranking_final", False))

    def _preencher_inscricoes_brutas(self, worksheet, inscricoes: list[Inscricao]) -> None:
        colunas_fixas = ["inscricao_id", "empresa_nome", "cnpj", "data_submissao"]
        colunas_dinamicas: list[str] = []
        for inscricao in inscricoes:
            for chave in inscricao.to_flat_dict().keys():
                if chave not in colunas_fixas and chave not in colunas_dinamicas:
                    colunas_dinamicas.append(chave)
        cabecalho = colunas_fixas + sorted(colunas_dinamicas)

        linhas = []
        for inscricao in inscricoes:
            dados = inscricao.to_flat_dict()
            linhas.append({coluna: dados.get(coluna) for coluna in cabecalho})

        self._escrever_tabela(worksheet, cabecalho, linhas)

    def _preencher_avaliacao_por_criterio(
        self,
        worksheet,
        ranking: list[ResultadoInscricao],
    ) -> None:
        cabecalho = [
            "inscricao_id",
            "empresa_nome",
            "categoria",
            "criterio_id",
            "criterio_nome",
            "resultado",
            "pontuacao",
            "pontuacao_maxima",
            "valor_observado",
            "justificativa",
            "revisao_humana_pendente",
            "contabilizar_na_nota",
        ]
        linhas = []
        for resultado in ranking:
            for criterio in resultado.resultados_criterios:
                linhas.append(criterio.to_dict(resultado.inscricao))
        self._escrever_tabela(worksheet, cabecalho, linhas)

    def _preencher_ranking_final(
        self,
        worksheet,
        ranking: list[ResultadoInscricao],
    ) -> None:
        criterios_pontuacao = self._coletar_criterios_pontuacao(ranking)
        criterios_revisao = self._coletar_criterios_revisao(ranking)
        cabecalho = ["ordem", "empresa_nome", "razao_social", "cnpj"]
        cabecalho += [self._coluna_pontuacao_criterio(criterio_id) for criterio_id, _ in criterios_pontuacao]
        cabecalho += ["pontuacao_total", "status_final", "motivo_status_final", "resumo_classificacao"]
        cabecalho += [self._coluna_revisao_criterio(criterio_id) for criterio_id, _ in criterios_revisao]

        labels = {
            "ordem": "Posicao",
            "empresa_nome": "Empresa",
            "razao_social": "Razao social",
            "cnpj": "CNPJ",
            "pontuacao_total": "Nota total",
            "status_final": "Status",
            "motivo_status_final": "Observacao",
            "resumo_classificacao": "Resumo da classificacao",
        }
        for criterio_id, criterio_nome in criterios_pontuacao:
            labels[self._coluna_pontuacao_criterio(criterio_id)] = (
                f"Pontos - {self._formatar_nome_criterio(criterio_nome)}"
            )
        for criterio_id, criterio_nome in criterios_revisao:
            labels[self._coluna_revisao_criterio(criterio_id)] = (
                f"Revisao humana - {self._formatar_nome_criterio(criterio_nome)}"
            )

        linhas = []
        for ordem, resultado in enumerate(ranking, start=1):
            linha = {
                "ordem": ordem,
                "empresa_nome": resultado.inscricao.empresa_nome,
                "razao_social": resultado.inscricao.obter_campo("razao_social"),
                "cnpj": resultado.inscricao.cnpj,
                "status_final": resultado.status_final,
                "motivo_status_final": resultado.motivo_status_final,
                "resumo_classificacao": self._montar_resumo_classificacao(resultado),
            }
            for criterio_id, criterio_nome in criterios_pontuacao:
                coluna = self._coluna_pontuacao_criterio(criterio_id)
                resultado_criterio = resultado.obter_resultado_criterio(criterio_id)
                linha[coluna] = resultado_criterio.pontuacao if resultado_criterio else None
            linha["pontuacao_total"] = resultado.pontuacao_total
            for criterio_id, _ in criterios_revisao:
                coluna = self._coluna_revisao_criterio(criterio_id)
                resultado_criterio = resultado.obter_resultado_criterio(criterio_id)
                linha[coluna] = self._texto_revisao_criterio(resultado_criterio)
            linhas.append(linha)

        self._escrever_tabela(worksheet, cabecalho, linhas, labels=labels)
        self._aplicar_destaque_top_20(worksheet, cabecalho)

    def _preencher_pendencias_revisao(
        self,
        worksheet,
        ranking: list[ResultadoInscricao],
    ) -> None:
        cabecalho = [
            "inscricao_id",
            "empresa_nome",
            "criterio_id",
            "criterio_nome",
            "pontuacao_sugerida",
            "valor_observado",
            "justificativa",
            "status_final_inscricao",
        ]
        linhas: list[dict[str, Any]] = []
        for resultado in ranking:
            for criterio in resultado.resultados_criterios:
                if criterio.revisao_humana_pendente:
                    linhas.append(
                        {
                            "inscricao_id": resultado.inscricao.inscricao_id,
                            "empresa_nome": resultado.inscricao.empresa_nome,
                            "criterio_id": criterio.criterio_id,
                            "criterio_nome": criterio.criterio_nome,
                            "pontuacao_sugerida": criterio.pontuacao,
                            "valor_observado": criterio.valor_observado,
                            "justificativa": criterio.justificativa,
                            "status_final_inscricao": resultado.status_final,
                        }
                    )
        self._escrever_tabela(worksheet, cabecalho, linhas)

    def _escrever_tabela(
        self,
        worksheet,
        cabecalho: list[str],
        linhas: list[dict[str, Any]],
        *,
        labels: dict[str, str] | None = None,
    ) -> None:
        worksheet.append([labels.get(coluna, coluna) if labels else coluna for coluna in cabecalho])
        self._estilizar_cabecalho(worksheet)
        worksheet.freeze_panes = "A2"

        for linha in linhas:
            worksheet.append([linha.get(coluna) for coluna in cabecalho])

        worksheet.auto_filter.ref = worksheet.dimensions
        self._estilizar_corpo(worksheet, cabecalho)
        self._ajustar_largura_colunas(worksheet, cabecalho)

    def _configurar_aparencia_planilha(self, worksheet) -> None:
        worksheet.sheet_view.showGridLines = False

    def _aplicar_cores_abas(self, workbook: Workbook) -> None:
        cores = {
            "inscricoes_brutas": COR_AZUL,
            "avaliacao_por_criterio": "FFCF24",
            "ranking_final": "2FD126",
            "pendencias_revisao": "FF914D",
        }
        for worksheet in workbook.worksheets:
            cor = cores.get(worksheet.title)
            if cor:
                worksheet.sheet_properties.tabColor = cor

    def _estilizar_cabecalho(self, worksheet) -> None:
        preenchimento = PatternFill(fill_type="solid", fgColor=COR_AZUL_ESCURO)
        borda = Border(
            left=Side(style="thin", color=COR_BORDA),
            right=Side(style="thin", color=COR_BORDA),
            top=Side(style="thin", color=COR_BORDA),
            bottom=Side(style="thin", color=COR_BORDA),
        )
        for celula in worksheet[1]:
            celula.font = Font(bold=True, color="FFFFFF")
            celula.fill = preenchimento
            celula.border = borda
            celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[1].height = 28

    def _estilizar_corpo(self, worksheet, cabecalho: list[str]) -> None:
        borda = Border(
            left=Side(style="thin", color=COR_BORDA),
            right=Side(style="thin", color=COR_BORDA),
            top=Side(style="thin", color=COR_BORDA),
            bottom=Side(style="thin", color=COR_BORDA),
        )
        preenchimento_impar = PatternFill(fill_type="solid", fgColor="FFFFFF")
        preenchimento_par = PatternFill(fill_type="solid", fgColor=COR_AZUL_CLARO_2)

        for indice_linha in range(2, worksheet.max_row + 1):
            for indice_coluna, nome_coluna in enumerate(cabecalho, start=1):
                celula = worksheet.cell(row=indice_linha, column=indice_coluna)
                celula.border = borda
                celula.fill = preenchimento_par if indice_linha % 2 == 0 else preenchimento_impar
                celula.alignment = self._alinhamento_coluna(nome_coluna)
                self._formatar_celula(celula, nome_coluna)

    def _formatar_celula(self, celula, nome_coluna: str) -> None:
        valor = celula.value

        if self._eh_coluna_pontuacao_criterio(nome_coluna) and isinstance(valor, (int, float)):
            celula.number_format = "0.00"
            celula.fill = PatternFill(fill_type="solid", fgColor=COR_AMARELO)
            celula.font = Font(color=COR_AMARELO_TEXTO, bold=True)
        elif self._eh_coluna_revisao_criterio(nome_coluna):
            if valor:
                celula.fill = PatternFill(fill_type="solid", fgColor=COR_AMARELO)
                celula.font = Font(color=COR_AMARELO_TEXTO, bold=False)
        elif nome_coluna in COLUNAS_NUMERICAS_DECIMAIS and isinstance(valor, (int, float)):
            celula.number_format = "0.00"
        elif nome_coluna in COLUNAS_NUMERICAS_INTEIRAS and isinstance(valor, (int, float)):
            celula.number_format = "0"
        elif nome_coluna in COLUNAS_DATA and valor:
            celula.number_format = "dd/mm/yyyy hh:mm"

        if nome_coluna in COLUNAS_STATUS:
            self._aplicar_estilo_status(celula)
        elif nome_coluna in COLUNAS_BOOLEANAS:
            self._aplicar_estilo_booleano(celula)
        elif nome_coluna in COLUNAS_NUMERICAS_DECIMAIS and isinstance(valor, (int, float)):
            celula.fill = PatternFill(fill_type="solid", fgColor=COR_AMARELO)
            celula.font = Font(color=COR_AMARELO_TEXTO, bold=True)

        if self._parece_url(valor):
            link = str(valor)
            if link.lower().startswith("www."):
                link = f"https://{link}"
            celula.hyperlink = link
            celula.font = Font(color=COR_AZUL, underline="single")

    def _aplicar_estilo_status(self, celula) -> None:
        valor = "" if celula.value is None else str(celula.value).strip().lower()
        mapa = {
            "classificada": (COR_VERDE, COR_VERDE_TEXTO),
            "classificada_com_revisao_pendente": (COR_AMARELO, COR_AMARELO_TEXTO),
            "aguardando_analise_de_demanda": (COR_AMARELO, COR_AMARELO_TEXTO),
            "eliminada": (COR_VERMELHO, COR_VERMELHO_TEXTO),
            "abaixo_do_corte": (COR_VERMELHO, COR_VERMELHO_TEXTO),
            "nao_selecionada_por_demanda": (COR_VERMELHO, COR_VERMELHO_TEXTO),
            "aprovado": (COR_VERDE, COR_VERDE_TEXTO),
            "pontuado": (COR_VERDE, COR_VERDE_TEXTO),
            "sugerido": (COR_AMARELO, COR_AMARELO_TEXTO),
            "pendente_revisao": (COR_AMARELO, COR_AMARELO_TEXTO),
            "reprovado": (COR_VERMELHO, COR_VERMELHO_TEXTO),
            "nao_pontuado": (COR_CINZA, COR_CINZA_TEXTO),
        }
        cor_fundo, cor_texto = mapa.get(valor, (COR_CINZA, COR_CINZA_TEXTO))
        celula.fill = PatternFill(fill_type="solid", fgColor=cor_fundo)
        celula.font = Font(color=cor_texto, bold=True)

    def _aplicar_estilo_booleano(self, celula) -> None:
        valor = celula.value
        if valor is True:
            celula.fill = PatternFill(fill_type="solid", fgColor=COR_VERDE)
            celula.font = Font(color=COR_VERDE_TEXTO, bold=True)
        elif valor is False:
            celula.fill = PatternFill(fill_type="solid", fgColor=COR_VERMELHO)
            celula.font = Font(color=COR_VERMELHO_TEXTO, bold=True)
        elif valor is None:
            celula.fill = PatternFill(fill_type="solid", fgColor=COR_AMARELO)
            celula.font = Font(color=COR_AMARELO_TEXTO, bold=True)

    def _alinhamento_coluna(self, nome_coluna: str) -> Alignment:
        if nome_coluna in COLUNAS_STATUS or nome_coluna in COLUNAS_BOOLEANAS:
            return Alignment(horizontal="center", vertical="center", wrap_text=True)
        if (
            nome_coluna in COLUNAS_NUMERICAS_DECIMAIS
            or nome_coluna in COLUNAS_NUMERICAS_INTEIRAS
            or self._eh_coluna_pontuacao_criterio(nome_coluna)
        ):
            return Alignment(horizontal="center", vertical="center")
        if self._eh_coluna_revisao_criterio(nome_coluna):
            return Alignment(horizontal="left", vertical="top", wrap_text=True)
        return Alignment(horizontal="left", vertical="top", wrap_text=nome_coluna in COLUNAS_LARGAS)

    def _ajustar_largura_colunas(self, worksheet, cabecalho: list[str]) -> None:
        for indice_coluna, coluna in enumerate(worksheet.columns, start=1):
            valores = ["" if celula.value is None else str(celula.value) for celula in coluna]
            nome_coluna = cabecalho[indice_coluna - 1] if indice_coluna - 1 < len(cabecalho) else ""
            if self._eh_coluna_pontuacao_criterio(nome_coluna):
                limite = 24
            elif self._eh_coluna_revisao_criterio(nome_coluna):
                limite = 80
            else:
                limite = 80 if nome_coluna in COLUNAS_LARGAS else 28
            largura = min(max(len(valor) for valor in valores) + 2, limite) if valores else 12
            largura = max(largura, 14 if nome_coluna in COLUNAS_STATUS else 12)
            worksheet.column_dimensions[get_column_letter(indice_coluna)].width = largura

    def _parece_url(self, valor: Any) -> bool:
        if valor is None:
            return False
        texto = str(valor).strip()
        if not texto:
            return False
        return re.match(r"^(https?://|www\.)", texto, flags=re.IGNORECASE) is not None

    def _eh_coluna_pontuacao_criterio(self, nome_coluna: str) -> bool:
        return nome_coluna.startswith("pontos__")

    def _eh_coluna_revisao_criterio(self, nome_coluna: str) -> bool:
        return nome_coluna.startswith("revisao__")

    def _coletar_criterios_pontuacao(
        self,
        ranking: list[ResultadoInscricao],
    ) -> list[tuple[str, str]]:
        criterios: list[tuple[str, str]] = []
        vistos: set[str] = set()
        for resultado in ranking:
            for criterio in resultado.resultados_criterios:
                if criterio.categoria != "pontuacao" or criterio.criterio_id in vistos:
                    continue
                vistos.add(criterio.criterio_id)
                criterios.append((criterio.criterio_id, criterio.criterio_nome))
        return criterios

    def _coletar_criterios_revisao(
        self,
        ranking: list[ResultadoInscricao],
    ) -> list[tuple[str, str]]:
        criterios: list[tuple[str, str]] = []
        vistos: set[str] = set()
        for resultado in ranking:
            for criterio in resultado.resultados_criterios:
                if not criterio.revisao_humana_pendente or criterio.criterio_id in vistos:
                    continue
                vistos.add(criterio.criterio_id)
                criterios.append((criterio.criterio_id, criterio.criterio_nome))
        return criterios

    def _coluna_pontuacao_criterio(self, criterio_id: str) -> str:
        return f"pontos__{criterio_id}"

    def _coluna_revisao_criterio(self, criterio_id: str) -> str:
        return f"revisao__{criterio_id}"

    def _formatar_nome_criterio(self, criterio_nome: str) -> str:
        texto = str(criterio_nome or "").strip().replace("_", " ")
        if not texto:
            return "criterio"
        return texto[0].upper() + texto[1:]

    def _montar_resumo_classificacao(self, resultado: ResultadoInscricao) -> str:
        partes: list[str] = []
        for criterio in resultado.resultados_criterios:
            if criterio.categoria == "pontuacao":
                prefixo = f"{criterio.criterio_nome}: {criterio.pontuacao:.2f}"
                if criterio.pontuacao_maxima is not None:
                    prefixo += f"/{criterio.pontuacao_maxima:.2f}"
            else:
                prefixo = f"{criterio.criterio_nome}: {criterio.resultado}"

            justificativa = (criterio.justificativa or "").strip()
            if justificativa:
                partes.append(f"{prefixo}. {justificativa}")
            else:
                partes.append(prefixo)

        return "\n".join(partes)

    def _texto_revisao_criterio(self, resultado_criterio) -> str | None:
        if resultado_criterio is None or not resultado_criterio.revisao_humana_pendente:
            return None

        justificativa = (resultado_criterio.justificativa or "").strip()
        if justificativa:
            return justificativa
        return "Criterio encaminhado para revisao humana."

    def _aplicar_destaque_top_20(self, worksheet, cabecalho: list[str]) -> None:
        limite = min(worksheet.max_row, 21)
        colunas_neutras = {
            "ordem",
            "empresa_nome",
            "razao_social",
            "cnpj",
            "pontuacao_total",
            "motivo_status_final",
            "resumo_classificacao",
        }

        for indice_linha in range(2, limite + 1):
            preenchimento_posicao, preenchimento_top, fonte_posicao = self._estilo_top_20(indice_linha - 1)
            for indice_coluna, nome_coluna in enumerate(cabecalho, start=1):
                celula = worksheet.cell(row=indice_linha, column=indice_coluna)
                if nome_coluna == "ordem":
                    celula.fill = preenchimento_posicao
                    celula.font = fonte_posicao
                elif (
                    nome_coluna in colunas_neutras
                    and nome_coluna not in COLUNAS_STATUS
                    and not self._eh_coluna_pontuacao_criterio(nome_coluna)
                ):
                    celula.fill = preenchimento_top

    def _estilo_top_20(self, posicao: int) -> tuple[PatternFill, PatternFill, Font]:
        if posicao <= 5:
            return (
                PatternFill(fill_type="solid", fgColor="1F8F43"),
                PatternFill(fill_type="solid", fgColor="E3F6E7"),
                Font(color="FFFFFF", bold=True),
            )
        if posicao <= 10:
            return (
                PatternFill(fill_type="solid", fgColor="3A78D6"),
                PatternFill(fill_type="solid", fgColor="E8F0FF"),
                Font(color="FFFFFF", bold=True),
            )
        return (
            PatternFill(fill_type="solid", fgColor="FFCF24"),
            PatternFill(fill_type="solid", fgColor="FFF7D8"),
            Font(color="6B4A00", bold=True),
        )

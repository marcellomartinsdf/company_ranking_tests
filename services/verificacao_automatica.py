from __future__ import annotations

import base64
import csv
import json
import mimetypes
import os
import re
import ssl
from dataclasses import dataclass
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote_plus, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from pypdf import PdfReader

from models import Inscricao
from models.inscricao import normalizar_chave, normalizar_texto


LEGAL_ENTITY_TOKENS = {
    "a",
    "agencia",
    "brasil",
    "comercio",
    "consultoria",
    "da",
    "de",
    "do",
    "dos",
    "e",
    "em",
    "exportacoes",
    "industria",
    "internacional",
    "ltda",
    "me",
    "mei",
    "s",
    "sa",
    "servicos",
}
PLACEHOLDER_VALUES = {"na", "n/a", "nenhum", "nenhuma", "nao possui", "xxxxx", "xxxxxxxxx"}
SOCIAL_DOMAINS = {
    "facebook.com",
    "instagram.com",
    "linkedin.com",
    "tiktok.com",
    "twitter.com",
    "x.com",
    "youtube.com",
}
LANGUAGE_STOPWORDS = {
    "en": {"and", "for", "from", "home", "products", "solutions", "the", "to", "with", "your"},
    "es": {"con", "contacto", "de", "en", "inicio", "nosotros", "para", "productos", "soluciones", "y"},
    "pt": {"com", "contato", "empresa", "inicio", "para", "produtos", "quem", "somos", "solucoes", "sobre"},
}


@dataclass
class RecursoRemoto:
    url: str
    final_url: str | None = None
    status_code: int | None = None
    content_type: str | None = None
    texto: str | None = None
    conteudo: bytes | None = None
    html_lang: str | None = None
    title: str | None = None
    erro: str | None = None


@dataclass
class ConsultaCNPJOficial:
    cnpj: str
    razao_social: str | None = None
    nome_fantasia: str | None = None
    situacao_cadastral: str | None = None
    fonte: str | None = None
    encontrado: bool = True
    erro: str | None = None


class _ExtratorHTML(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._em_title = False
        self._textos: list[str] = []
        self.title = ""
        self.html_lang: str | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        atributos = dict(attrs)
        if tag == "title":
            self._em_title = True
        if tag == "html":
            self.html_lang = atributos.get("lang")

    def handle_endtag(self, tag: str) -> None:
        if tag == "title":
            self._em_title = False

    def handle_data(self, data: str) -> None:
        texto = data.strip()
        if not texto:
            return
        if self._em_title:
            self.title += f" {texto}"
        self._textos.append(texto)

    @property
    def texto(self) -> str:
        return " ".join(self._textos)


class ClienteOpenAI:
    def __init__(self, config: dict[str, Any] | None = None) -> None:
        self.config = config or {}
        self.habilitado = bool(self.config.get("habilitado", True))
        self.api_key = os.environ.get("OPENAI_API_KEY")
        self.base_url = self.config.get("base_url", "https://api.openai.com/v1/responses")
        self.model = self.config.get("model", "gpt-4.1-mini")
        self.timeout = float(self.config.get("timeout_segundos", 25))

    @property
    def disponivel(self) -> bool:
        return self.habilitado and bool(self.api_key)

    def analisar_canal(
        self,
        *,
        empresa_nome: str,
        razao_social: str | None,
        url: str,
        titulo: str | None,
        html_lang: str | None,
        texto: str,
    ) -> dict[str, Any] | None:
        if not self.disponivel:
            return None

        prompt = (
            "Responda em JSON. Analise se o canal informado parece oficial da empresa, "
            "se ha conteudo em ingles e/ou espanhol e se o material aparenta ser voltado "
            "ao comprador estrangeiro.\n"
            f"Empresa: {empresa_nome}\n"
            f"Razao social: {razao_social or ''}\n"
            f"URL: {url}\n"
            f"Title: {titulo or ''}\n"
            f"HTML lang: {html_lang or ''}\n"
            f"Trecho extraido: {texto[:4000]}"
        )
        schema = {
            "type": "object",
            "properties": {
                "official_channel": {"type": ["boolean", "null"]},
                "languages": {"type": "array", "items": {"type": "string"}},
                "foreign_buyer_ready": {"type": ["boolean", "null"]},
                "explanation": {"type": "string"},
            },
            "required": ["official_channel", "languages", "foreign_buyer_ready", "explanation"],
            "additionalProperties": False,
        }
        return self._solicitar_json(prompt=prompt, schema=schema)

    def analisar_cartao_cnpj(
        self,
        *,
        empresa_nome: str,
        razao_social: str | None,
        cnpj: str,
        recurso: RecursoRemoto,
        texto_extraido: str | None,
    ) -> dict[str, Any] | None:
        if not self.disponivel:
            return None

        conteudo: list[dict[str, Any]] = [
            {
                "type": "input_text",
                "text": (
                    "Responda em JSON. Verifique se o documento parece ser um cartao CNPJ "
                    "ou comprovante cadastral e se ele confere com a empresa informada.\n"
                    f"Empresa: {empresa_nome}\n"
                    f"Razao social: {razao_social or ''}\n"
                    f"CNPJ esperado: {cnpj}\n"
                    f"Texto parcial ja extraido: {(texto_extraido or '')[:3000]}"
                ),
            }
        ]

        if recurso.content_type and "pdf" in recurso.content_type and recurso.conteudo:
            conteudo.append(
                {
                    "type": "input_file",
                    "filename": "cartao-cnpj.pdf",
                    "file_data": (
                        "data:application/pdf;base64,"
                        f"{base64.b64encode(recurso.conteudo).decode('utf-8')}"
                    ),
                }
            )
        elif recurso.content_type and recurso.content_type.startswith("image/") and recurso.conteudo:
            data_url = (
                f"data:{recurso.content_type};base64,"
                f"{base64.b64encode(recurso.conteudo).decode('utf-8')}"
            )
            conteudo.append({"type": "input_image", "image_url": data_url})

        schema = {
            "type": "object",
            "properties": {
                "document_kind": {"type": "string"},
                "cnpj_match": {"type": ["boolean", "null"]},
                "company_match": {"type": ["boolean", "null"]},
                "explanation": {"type": "string"},
            },
            "required": ["document_kind", "cnpj_match", "company_match", "explanation"],
            "additionalProperties": False,
        }
        return self._solicitar_json(conteudo=conteudo, schema=schema)

    def _solicitar_json(
        self,
        *,
        prompt: str | None = None,
        conteudo: list[dict[str, Any]] | None = None,
        schema: dict[str, Any],
    ) -> dict[str, Any] | None:
        if not self.disponivel:
            return None

        itens_conteudo = conteudo or [{"type": "input_text", "text": prompt or "Responda em JSON."}]
        payload = {
            "model": self.model,
            "input": [{"role": "user", "content": itens_conteudo}],
            "text": {
                "format": {
                    "type": "json_schema",
                    "name": "verificacao",
                    "strict": True,
                    "schema": schema,
                }
            },
        }

        request = Request(
            self.base_url,
            data=json.dumps(payload).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        try:
            with urlopen(request, timeout=self.timeout) as resposta:
                dados = json.loads(resposta.read().decode("utf-8"))
        except Exception:
            return None

        texto_saida = dados.get("output_text")
        if not texto_saida:
            return None

        try:
            return json.loads(texto_saida)
        except json.JSONDecodeError:
            return None


class VerificacaoAutomaticaService:
    def __init__(self, regulamento: dict[str, Any]) -> None:
        self.config = regulamento.get("verificacoes_automaticas") or {}
        self.timeout_segundos = float(self.config.get("timeout_segundos", 12))
        self.user_agent = self.config.get(
            "user_agent",
            "Mozilla/5.0 (compatible; ApexBrasil-Ranqueamento/1.0; +https://apexbrasil.com.br)",
        )
        self.max_download_bytes = int(self.config.get("max_download_bytes", 10 * 1024 * 1024))
        self.max_recursos_remotos_por_execucao = self._normalizar_limite_recursos(
            self.config.get("max_recursos_remotos_por_execucao")
        )
        self._cache_recursos: dict[str, RecursoRemoto] = {}
        self._cache_cnpj_oficial: dict[str, ConsultaCNPJOficial] = {}
        self._cache_dataset_cnpj: dict[str, dict[str, Any]] | None = None
        self._recursos_remotos_processados = 0
        usar_openai = bool(self.config.get("usar_openai_quando_disponivel", True))
        config_openai = dict(self.config.get("openai") or {})
        config_openai["habilitado"] = usar_openai
        self.openai = ClienteOpenAI(config_openai)

    def enriquecer_inscricoes(self, inscricoes: list[Inscricao]) -> list[Inscricao]:
        for inscricao in inscricoes:
            self.enriquecer_inscricao(inscricao)
        return inscricoes

    def enriquecer_inscricao(self, inscricao: Inscricao) -> Inscricao:
        if not self.config:
            return inscricao

        self._derivar_status_financeiro(inscricao)
        self._verificar_cnpj(inscricao)
        self._verificar_canal_empresa(inscricao)
        self._verificar_canal_internacional(inscricao)
        self._verificar_cartao_cnpj(inscricao)
        return inscricao

    def _derivar_status_financeiro(self, inscricao: Inscricao) -> None:
        config = self.config.get("status_financeiro") or {}
        campo = config.get("campo", "status_financeiro")
        valor = inscricao.obter_campo(campo)
        texto = normalizar_texto(valor)

        if not texto:
            inscricao.definir_campo("possui_pendencia_financeira_apex", None)
            inscricao.definir_campo(
                "possui_pendencia_financeira_apex_justificativa",
                "Status financeiro ausente na exportacao do Dynamics.",
            )
            return

        valores_sem_pendencia = {
            normalizar_texto(item) for item in config.get("valores_sem_pendencia", ["adimplente"])
        }
        tokens_com_pendencia = [
            normalizar_texto(item)
            for item in config.get("tokens_com_pendencia", ["inadimpl", "pendenc", "debito"])
        ]

        if texto in valores_sem_pendencia:
            possui_pendencia: bool | None = False
            justificativa = f"Status financeiro '{valor}' nao indica pendencia financeira."
        elif any(token in texto for token in tokens_com_pendencia):
            possui_pendencia = True
            justificativa = f"Status financeiro '{valor}' indica pendencia com base na configuracao."
        else:
            possui_pendencia = None
            justificativa = (
                f"Status financeiro '{valor}' nao pode ser classificado automaticamente como "
                "adimplente ou pendente."
            )

        inscricao.definir_campo("possui_pendencia_financeira_apex", possui_pendencia)
        inscricao.definir_campo("possui_pendencia_financeira_apex_justificativa", justificativa)

    def _verificar_cnpj(self, inscricao: Inscricao) -> None:
        config = self.config.get("cnpj") or {}
        campo = config.get("campo", "cnpj")
        valor = inscricao.obter_campo(campo)
        cnpj = self._apenas_digitos(valor)
        nome_empresa = inscricao.obter_campo("razao_social") or inscricao.empresa_nome

        inscricao.definir_campo("cnpj_digitos", cnpj or None)
        inscricao.definir_campo("cnpj_valor_observado", valor)

        if not cnpj:
            inscricao.definir_campo("cnpj_valido", False)
            inscricao.definir_campo("cnpj_justificativa", "CNPJ ausente na inscricao.")
            return

        valido = self._validar_cnpj(cnpj)
        inscricao.definir_campo("cnpj_valido", valido)
        if valido:
            justificativa = (
                f"CNPJ {self._formatar_cnpj(cnpj)} passou na validacao estrutural dos digitos verificadores."
            )
        else:
            justificativa = (
                f"CNPJ {self._formatar_cnpj(cnpj)} nao passou na validacao estrutural."
            )
        inscricao.definir_campo("cnpj_justificativa", justificativa)
        inscricao.definir_campo("cnpj_nome_referencia", nome_empresa)
        self._verificar_cnpj_em_fonte_oficial(
            inscricao,
            cnpj=cnpj,
            cnpj_valido=valido,
            razao_social_local=inscricao.obter_campo("razao_social"),
        )

    def _verificar_cnpj_em_fonte_oficial(
        self,
        inscricao: Inscricao,
        *,
        cnpj: str,
        cnpj_valido: bool,
        razao_social_local: str | None,
    ) -> None:
        config = self.config.get("cnpj_consulta_oficial") or {}
        prefixo = config.get("campo_saida", "cnpj_empresa_confere_cadastro_oficial")
        empresa_nome = inscricao.empresa_nome
        cnpj_formatado = self._formatar_cnpj(cnpj) if cnpj else ""

        if not cnpj:
            self._registrar_resultado_verificacao(
                inscricao,
                prefixo=prefixo,
                valor=None,
                revisao_humana=False,
                valor_observado=cnpj_formatado or None,
                justificativa="Nao foi possivel consultar fonte oficial porque o CNPJ esta ausente.",
                campos_extras={
                    "cnpj_fonte_oficial_consulta": None,
                    "cnpj_fonte_oficial_resumo": None,
                    "cnpj_cadastro_oficial_razao_social": None,
                    "cnpj_cadastro_oficial_nome_fantasia": None,
                    "cnpj_cadastro_oficial_situacao_cadastral": None,
                },
            )
            return

        if not cnpj_valido:
            self._registrar_resultado_verificacao(
                inscricao,
                prefixo=prefixo,
                valor=False,
                revisao_humana=False,
                valor_observado=cnpj_formatado,
                justificativa=(
                    "Nao foi feita comparacao com cadastro oficial porque o CNPJ falhou "
                    "na validacao estrutural."
                ),
                campos_extras={
                    "cnpj_fonte_oficial_consulta": None,
                    "cnpj_fonte_oficial_resumo": cnpj_formatado,
                    "cnpj_cadastro_oficial_razao_social": None,
                    "cnpj_cadastro_oficial_nome_fantasia": None,
                    "cnpj_cadastro_oficial_situacao_cadastral": None,
                },
            )
            return

        if not config.get("habilitado", False):
            self._registrar_resultado_verificacao(
                inscricao,
                prefixo=prefixo,
                valor=None,
                revisao_humana=False,
                valor_observado=cnpj_formatado,
                justificativa=(
                    "Nenhuma fonte oficial de consulta de CNPJ esta configurada. "
                    "A verificacao automatica ficou restrita a estrutura do CNPJ e aos documentos enviados."
                ),
                campos_extras={
                    "cnpj_fonte_oficial_consulta": None,
                    "cnpj_fonte_oficial_resumo": cnpj_formatado,
                    "cnpj_cadastro_oficial_razao_social": None,
                    "cnpj_cadastro_oficial_nome_fantasia": None,
                    "cnpj_cadastro_oficial_situacao_cadastral": None,
                },
            )
            return

        consulta = self._consultar_cnpj_fonte_oficial(cnpj, config)
        if consulta.erro:
            self._registrar_resultado_verificacao(
                inscricao,
                prefixo=prefixo,
                valor=None,
                revisao_humana=True,
                valor_observado=cnpj_formatado,
                justificativa=f"Falha ao consultar fonte oficial do CNPJ: {consulta.erro}.",
                campos_extras={
                    "cnpj_fonte_oficial_consulta": consulta.fonte,
                    "cnpj_fonte_oficial_resumo": cnpj_formatado,
                    "cnpj_cadastro_oficial_razao_social": consulta.razao_social,
                    "cnpj_cadastro_oficial_nome_fantasia": consulta.nome_fantasia,
                    "cnpj_cadastro_oficial_situacao_cadastral": consulta.situacao_cadastral,
                },
            )
            return

        if not consulta.encontrado:
            self._registrar_resultado_verificacao(
                inscricao,
                prefixo=prefixo,
                valor=False,
                revisao_humana=False,
                valor_observado=cnpj_formatado,
                justificativa=(
                    "O CNPJ nao foi encontrado na fonte oficial configurada para consulta automatica."
                ),
                campos_extras={
                    "cnpj_fonte_oficial_consulta": consulta.fonte,
                    "cnpj_fonte_oficial_resumo": cnpj_formatado,
                    "cnpj_cadastro_oficial_razao_social": None,
                    "cnpj_cadastro_oficial_nome_fantasia": None,
                    "cnpj_cadastro_oficial_situacao_cadastral": None,
                },
            )
            return

        confere, revisao_humana, justificativa = self._comparar_empresa_com_cadastro_oficial(
            empresa_nome=empresa_nome,
            razao_social_local=razao_social_local,
            consulta=consulta,
        )
        resumo = self._montar_resumo_consulta_oficial(cnpj, consulta)
        self._registrar_resultado_verificacao(
            inscricao,
            prefixo=prefixo,
            valor=confere,
            revisao_humana=revisao_humana,
            valor_observado=resumo,
            justificativa=justificativa,
            campos_extras={
                "cnpj_fonte_oficial_consulta": consulta.fonte,
                "cnpj_fonte_oficial_resumo": resumo,
                "cnpj_cadastro_oficial_razao_social": consulta.razao_social,
                "cnpj_cadastro_oficial_nome_fantasia": consulta.nome_fantasia,
                "cnpj_cadastro_oficial_situacao_cadastral": consulta.situacao_cadastral,
            },
        )

    def _consultar_cnpj_fonte_oficial(
        self,
        cnpj: str,
        config: dict[str, Any],
    ) -> ConsultaCNPJOficial:
        if cnpj in self._cache_cnpj_oficial:
            return self._cache_cnpj_oficial[cnpj]

        fonte = normalizar_texto(config.get("fonte", "arquivo_local"))
        if fonte == "arquivo_local":
            consulta = self._consultar_cnpj_arquivo_local(cnpj, config)
        elif fonte == "http_api":
            consulta = self._consultar_cnpj_http_api(cnpj, config)
        else:
            consulta = ConsultaCNPJOficial(
                cnpj=cnpj,
                fonte=fonte or "desconhecida",
                erro=f"Fonte oficial de CNPJ nao suportada: {config.get('fonte')}.",
            )

        self._cache_cnpj_oficial[cnpj] = consulta
        return consulta

    def _consultar_cnpj_arquivo_local(
        self,
        cnpj: str,
        config: dict[str, Any],
    ) -> ConsultaCNPJOficial:
        caminho = self._expandir_variaveis_ambiente(config.get("arquivo"))
        if not caminho:
            return ConsultaCNPJOficial(
                cnpj=cnpj,
                fonte="arquivo_local",
                erro="Caminho do dataset local de CNPJ nao configurado.",
            )

        path = Path(caminho)
        if not path.exists():
            return ConsultaCNPJOficial(
                cnpj=cnpj,
                fonte=str(path),
                erro="Arquivo local de consulta de CNPJ nao encontrado.",
            )

        try:
            indice = self._carregar_dataset_cnpj(path, config)
        except Exception as exc:
            return ConsultaCNPJOficial(
                cnpj=cnpj,
                fonte=str(path),
                erro=f"Falha ao carregar dataset local de CNPJ: {exc}",
            )

        registro = indice.get(cnpj)
        if registro is None:
            return ConsultaCNPJOficial(cnpj=cnpj, fonte=str(path), encontrado=False)

        return ConsultaCNPJOficial(
            cnpj=cnpj,
            razao_social=self._buscar_valor_registro(registro, config.get("coluna_razao_social", "razao_social")),
            nome_fantasia=self._buscar_valor_registro(registro, config.get("coluna_nome_fantasia", "nome_fantasia")),
            situacao_cadastral=self._buscar_valor_registro(
                registro,
                config.get("coluna_situacao_cadastral", "situacao_cadastral"),
            ),
            fonte=str(path),
            encontrado=True,
        )

    def _consultar_cnpj_http_api(
        self,
        cnpj: str,
        config: dict[str, Any],
    ) -> ConsultaCNPJOficial:
        url_template = self._expandir_variaveis_ambiente(config.get("url_template"))
        if not url_template:
            return ConsultaCNPJOficial(
                cnpj=cnpj,
                fonte="http_api",
                erro="URL da API oficial de CNPJ nao configurada.",
            )

        url = url_template.format(cnpj=cnpj, cnpj_formatado=self._formatar_cnpj(cnpj))
        headers = self._montar_headers_http(config.get("headers"))
        metodo = str(config.get("metodo", "GET")).upper()
        request = Request(url, headers=headers, method=metodo)

        try:
            context = ssl.create_default_context()
            with urlopen(request, timeout=self.timeout_segundos, context=context) as resposta:
                payload = json.loads(resposta.read().decode("utf-8"))
        except HTTPError as exc:
            if exc.code == 404:
                return ConsultaCNPJOficial(cnpj=cnpj, fonte=url, encontrado=False)
            return ConsultaCNPJOficial(
                cnpj=cnpj,
                fonte=url,
                erro=f"HTTP {exc.code}",
            )
        except Exception as exc:
            return ConsultaCNPJOficial(
                cnpj=cnpj,
                fonte=url,
                erro=str(exc),
            )

        paths = config.get("response_path") or {}
        return ConsultaCNPJOficial(
            cnpj=self._apenas_digitos(self._extrair_valor_json(payload, paths.get("cnpj"))) or cnpj,
            razao_social=self._extrair_valor_json(payload, paths.get("razao_social")) if paths.get("razao_social") else None,
            nome_fantasia=self._extrair_valor_json(payload, paths.get("nome_fantasia")) if paths.get("nome_fantasia") else None,
            situacao_cadastral=self._extrair_valor_json(payload, paths.get("situacao_cadastral")) if paths.get("situacao_cadastral") else None,
            fonte=url,
            encontrado=True,
        )

    def _carregar_dataset_cnpj(
        self,
        path: Path,
        config: dict[str, Any],
    ) -> dict[str, dict[str, Any]]:
        if self._cache_dataset_cnpj is not None:
            return self._cache_dataset_cnpj

        coluna_cnpj = normalizar_texto(config.get("coluna_cnpj", "cnpj"))
        extensao = path.suffix.lower()
        indice: dict[str, dict[str, Any]] = {}

        if extensao == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as arquivo:
                leitor = csv.DictReader(arquivo)
                for linha in leitor:
                    registro = {normalizar_chave(chave): valor for chave, valor in linha.items() if chave}
                    cnpj_registro = self._apenas_digitos(registro.get(normalizar_chave(coluna_cnpj)))
                    if cnpj_registro:
                        indice[cnpj_registro] = registro
        elif extensao in {".xlsx", ".xlsm"}:
            workbook = load_workbook(filename=path, read_only=True, data_only=True)
            worksheet = workbook.active
            linhas = worksheet.iter_rows(values_only=True)
            cabecalho = next(linhas, None) or []
            chaves = [normalizar_chave(chave) for chave in cabecalho]
            for linha in linhas:
                registro = {
                    chaves[indice_coluna]: linha[indice_coluna]
                    for indice_coluna in range(min(len(chaves), len(linha)))
                    if indice_coluna < len(chaves) and chaves[indice_coluna]
                }
                cnpj_registro = self._apenas_digitos(registro.get(normalizar_chave(coluna_cnpj)))
                if cnpj_registro:
                    indice[cnpj_registro] = registro
            workbook.close()
        elif extensao == ".json":
            payload = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(payload, list):
                for item in payload:
                    if not isinstance(item, dict):
                        continue
                    registro = {normalizar_chave(chave): valor for chave, valor in item.items() if chave}
                    cnpj_registro = self._apenas_digitos(registro.get(normalizar_chave(coluna_cnpj)))
                    if cnpj_registro:
                        indice[cnpj_registro] = registro
        else:
            raise ValueError("Formato do dataset local de CNPJ nao suportado. Use CSV, XLSX ou JSON.")

        self._cache_dataset_cnpj = indice
        return indice

    def _comparar_empresa_com_cadastro_oficial(
        self,
        *,
        empresa_nome: str,
        razao_social_local: str | None,
        consulta: ConsultaCNPJOficial,
    ) -> tuple[bool | None, bool, str]:
        nomes_locais = [nome for nome in [empresa_nome, razao_social_local] if self._tem_valor(nome)]
        nomes_oficiais = [
            nome
            for nome in [consulta.razao_social, consulta.nome_fantasia]
            if self._tem_valor(nome)
        ]

        if not nomes_oficiais:
            return (
                None,
                True,
                "A fonte oficial retornou o CNPJ, mas sem razao social ou nome fantasia suficientes para confronto.",
            )

        for nome_local in nomes_locais:
            for nome_oficial in nomes_oficiais:
                if normalizar_texto(nome_local) == normalizar_texto(nome_oficial):
                    return (
                        True,
                        False,
                        "O nome da empresa confere exatamente com o cadastro oficial consultado.",
                    )

        tokens_locais = set(self._tokens_empresa(*nomes_locais))
        tokens_oficiais = set(self._tokens_empresa(*nomes_oficiais))
        intersecao = tokens_locais.intersection(tokens_oficiais)

        if intersecao:
            proporcao_local = len(intersecao) / max(len(tokens_locais), 1)
            proporcao_oficial = len(intersecao) / max(len(tokens_oficiais), 1)

            if len(intersecao) >= 2 and (proporcao_local >= 0.5 or proporcao_oficial >= 0.5):
                return (
                    True,
                    True,
                    "O cadastro oficial apresenta forte semelhanca textual com a empresa informada, "
                    f"mas vale revisao humana. Tokens em comum: {', '.join(sorted(intersecao))}.",
                )

            return (
                False,
                True,
                "Foi encontrada apenas semelhanca parcial entre a empresa informada e o cadastro oficial. "
                f"Tokens em comum: {', '.join(sorted(intersecao))}.",
            )

        resumo = self._montar_resumo_consulta_oficial(consulta.cnpj, consulta)
        return (
            False,
            False,
            "A empresa informada nao corresponde ao nome empresarial ou nome fantasia retornado "
            f"pela fonte oficial consultada: {resumo}.",
        )

    def _montar_resumo_consulta_oficial(
        self,
        cnpj: str,
        consulta: ConsultaCNPJOficial,
    ) -> str:
        partes = [self._formatar_cnpj(cnpj)]
        if self._tem_valor(consulta.razao_social):
            partes.append(str(consulta.razao_social))
        if self._tem_valor(consulta.nome_fantasia):
            partes.append(str(consulta.nome_fantasia))
        if self._tem_valor(consulta.situacao_cadastral):
            partes.append(f"Situacao: {consulta.situacao_cadastral}")
        return " | ".join(partes)

    def _buscar_valor_registro(self, registro: dict[str, Any], coluna: str) -> Any:
        return registro.get(normalizar_chave(coluna))

    def _montar_headers_http(self, headers: Any) -> dict[str, str]:
        if not isinstance(headers, dict):
            return {}
        resultado: dict[str, str] = {}
        for chave, valor in headers.items():
            if valor is None:
                continue
            resultado[str(chave)] = self._expandir_variaveis_ambiente(str(valor))
        return resultado

    def _expandir_variaveis_ambiente(self, valor: Any) -> str:
        if valor is None:
            return ""
        texto = str(valor)
        for variavel in re.findall(r"\$\{([^}]+)\}", texto):
            texto = texto.replace(f"${{{variavel}}}", os.environ.get(variavel, ""))
        return texto

    def _extrair_valor_json(self, payload: Any, caminho: str | None) -> Any:
        if caminho is None or caminho == "":
            return None
        atual = payload
        for parte in str(caminho).split("."):
            if isinstance(atual, dict):
                atual = atual.get(parte)
            elif isinstance(atual, list) and parte.isdigit():
                indice = int(parte)
                atual = atual[indice] if 0 <= indice < len(atual) else None
            else:
                return None
            if atual is None:
                return None
        return atual

    def _verificar_canal_empresa(self, inscricao: Inscricao) -> None:
        config = self.config.get("website_social") or {}
        campo = config.get("campo", "website_empresa")
        url_original = inscricao.obter_campo(campo)
        prefixo = config.get("campo_saida", "website_ou_rede_social_verificado")
        url_normalizada = self._normalizar_url_publica(url_original)

        if not self._tem_valor(url_original):
            self._registrar_resultado_verificacao(
                inscricao,
                prefixo=prefixo,
                valor=False,
                revisao_humana=False,
                valor_observado=url_original,
                justificativa="Empresa nao informou website ou rede social.",
                campos_extras={"website_ou_rede_social_url_normalizada": None},
            )
            return

        if url_normalizada is None:
            self._registrar_resultado_verificacao(
                inscricao,
                prefixo=prefixo,
                valor=None,
                revisao_humana=True,
                valor_observado=url_original,
                justificativa="Link informado nao parece um URL publico valido para verificacao automatica.",
                campos_extras={"website_ou_rede_social_url_normalizada": None},
            )
            return

        recurso = self._obter_recurso(url_normalizada)
        resultado = self._avaliar_canal(inscricao, url_normalizada, recurso)
        self._registrar_resultado_verificacao(
            inscricao,
            prefixo=prefixo,
            valor=resultado["valor"],
            revisao_humana=resultado["revisao_humana"],
            valor_observado=url_original,
            justificativa=resultado["justificativa"],
            campos_extras={
                "website_ou_rede_social_url_normalizada": url_normalizada,
                "website_ou_rede_social_tipo_canal": resultado["tipo_canal"],
            },
        )

    def _verificar_canal_internacional(self, inscricao: Inscricao) -> None:
        config = self.config.get("website_internacional") or {}
        campo_link = config.get("campo_link", "website_internacional_link")
        campo_resposta = config.get("campo_resposta", "website_internacional_resposta")
        prefixo = config.get("campo_saida", "website_internacional_verificado")
        resposta_declarada = inscricao.obter_campo(campo_resposta)
        link_informado = inscricao.obter_campo(campo_link) or inscricao.obter_campo("website_empresa")
        url_normalizada = self._normalizar_url_publica(link_informado)

        if not self._tem_valor(link_informado):
            valor = False if resposta_declarada in (False, None) else None
            revisao = resposta_declarada not in (False, None)
            justificativa = (
                "Nao ha link de website internacional para validar a declaracao."
                if revisao
                else "Empresa nao informou website internacional."
            )
            self._registrar_resultado_verificacao(
                inscricao,
                prefixo=prefixo,
                valor=valor,
                revisao_humana=revisao,
                valor_observado=link_informado,
                justificativa=justificativa,
                campos_extras={"website_internacional_url_normalizada": None},
            )
            return

        if url_normalizada is None:
            self._registrar_resultado_verificacao(
                inscricao,
                prefixo=prefixo,
                valor=None,
                revisao_humana=True,
                valor_observado=link_informado,
                justificativa="Link informado para vendas internacionais nao pode ser normalizado automaticamente.",
                campos_extras={"website_internacional_url_normalizada": None},
            )
            return

        recurso = self._obter_recurso(url_normalizada)
        idiomas_detectados = self._detectar_idiomas(recurso)
        possui_idioma_estrangeiro = any(idioma in {"en", "es"} for idioma in idiomas_detectados)
        revisao_humana = False
        justificativas = []

        if recurso.erro:
            justificativas.append(f"Falha ao acessar o canal informado: {recurso.erro}.")
            revisao_humana = True
            valor: bool | None = None
        else:
            valor = possui_idioma_estrangeiro
            if possui_idioma_estrangeiro:
                justificativas.append(
                    "Conteudo do canal apresenta indicios de ingles e/ou espanhol."
                )
            else:
                justificativas.append("Nao foram encontrados indicios suficientes de ingles ou espanhol.")

        if idiomas_detectados:
            justificativas.append(
                f"Idiomas detectados automaticamente: {', '.join(idiomas_detectados)}."
            )

        if resposta_declarada is True and valor is False:
            revisao_humana = True
            justificativas.append(
                "Declaracao da empresa foi positiva, mas a heuristica nao confirmou o idioma estrangeiro."
            )

        if self.openai.disponivel and recurso.texto and (valor is False or revisao_humana):
            parecer_ia = self.openai.analisar_canal(
                empresa_nome=inscricao.empresa_nome,
                razao_social=inscricao.obter_campo("razao_social"),
                url=url_normalizada,
                titulo=recurso.title,
                html_lang=recurso.html_lang,
                texto=recurso.texto[:6000],
            )
            if parecer_ia:
                idiomas_ia = [
                    normalizar_texto(item)
                    for item in parecer_ia.get("languages", [])
                    if normalizar_texto(item) in {"en", "es", "pt"}
                ]
                if any(idioma in {"en", "es"} for idioma in idiomas_ia):
                    valor = True
                if parecer_ia.get("foreign_buyer_ready") is True:
                    valor = True
                    revisao_humana = True
                justificativas.append(f"IA: {parecer_ia.get('explanation', '').strip()}")
                if idiomas_ia:
                    idiomas_detectados = sorted(set(idiomas_detectados + idiomas_ia))

        self._registrar_resultado_verificacao(
            inscricao,
            prefixo=prefixo,
            valor=valor,
            revisao_humana=revisao_humana,
            valor_observado=link_informado,
            justificativa=" ".join(justificativas).strip(),
            campos_extras={
                "website_internacional_url_normalizada": url_normalizada,
                "website_internacional_idiomas_detectados": ", ".join(idiomas_detectados),
            },
        )

    def _verificar_cartao_cnpj(self, inscricao: Inscricao) -> None:
        config = self.config.get("cartao_cnpj") or {}
        campos_link = config.get("campos_link", ["cartao_cnpj_link", "cartao_cnpj_regional_link"])
        prefixo = config.get("campo_saida", "cartao_cnpj_verificado")
        links = [inscricao.obter_campo(campo) for campo in campos_link]
        links_validos = [link for link in links if self._tem_valor(link)]
        cnpj = self._apenas_digitos(inscricao.cnpj)
        razao_social = inscricao.obter_campo("razao_social") or inscricao.empresa_nome

        if not links_validos:
            self._registrar_resultado_verificacao(
                inscricao,
                prefixo=prefixo,
                valor=False,
                revisao_humana=False,
                valor_observado=None,
                justificativa="Nao foi informado link para o cartao CNPJ.",
                campos_extras={
                    "cartao_cnpj_cnpj_confere": False,
                    "cartao_cnpj_razao_social_confere": False,
                    "cartao_cnpj_fonte": None,
                },
            )
            return

        link_escolhido = links_validos[0]
        link_normalizado = self._normalizar_link_documento(link_escolhido)
        if link_normalizado is None:
            self._registrar_resultado_verificacao(
                inscricao,
                prefixo=prefixo,
                valor=None,
                revisao_humana=True,
                valor_observado=link_escolhido,
                justificativa="Link do cartao CNPJ nao pode ser lido automaticamente no formato atual.",
                campos_extras={
                    "cartao_cnpj_cnpj_confere": None,
                    "cartao_cnpj_razao_social_confere": None,
                    "cartao_cnpj_fonte": link_escolhido,
                },
            )
            return

        recurso = self._obter_recurso(link_normalizado)
        texto_extraido = self._extrair_texto_documento(recurso)
        cnpj_confere = False
        razao_confere = False
        revisao_humana = False
        justificativas = []

        if recurso.erro:
            revisao_humana = True
            justificativas.append(f"Falha ao acessar o documento: {recurso.erro}.")
            valor: bool | None = None
        else:
            valor = False
            if texto_extraido:
                cnpj_confere = bool(cnpj and cnpj in self._apenas_digitos(texto_extraido))
                razao_confere = self._texto_corresponde_empresa(texto_extraido, inscricao.empresa_nome, razao_social)
                valor = cnpj_confere and razao_confere
                if valor:
                    justificativas.append(
                        "Texto extraido do documento contem o CNPJ e referencia compativel com a empresa."
                    )
                else:
                    justificativas.append(
                        "Documento foi lido, mas nao confirmou simultaneamente o CNPJ e a razao social."
                    )
                    revisao_humana = cnpj_confere or razao_confere
            else:
                revisao_humana = True
                valor = None
                justificativas.append(
                    "Nao foi possivel extrair texto util do documento para confrontar CNPJ e razao social."
                )

        if self.openai.disponivel and recurso.conteudo and (valor is None or not valor):
            parecer_ia = self.openai.analisar_cartao_cnpj(
                empresa_nome=inscricao.empresa_nome,
                razao_social=razao_social,
                cnpj=cnpj or "",
                recurso=recurso,
                texto_extraido=texto_extraido,
            )
            if parecer_ia:
                justificativas.append(f"IA: {parecer_ia.get('explanation', '').strip()}")
                if parecer_ia.get("cnpj_match") is True:
                    cnpj_confere = True
                if parecer_ia.get("company_match") is True:
                    razao_confere = True
                if parecer_ia.get("document_kind") == "cartao_cnpj":
                    revisao_humana = revisao_humana or (not (cnpj_confere and razao_confere))
                valor = cnpj_confere and razao_confere

        self._registrar_resultado_verificacao(
            inscricao,
            prefixo=prefixo,
            valor=valor,
            revisao_humana=revisao_humana,
            valor_observado=link_escolhido,
            justificativa=" ".join(justificativas).strip(),
            campos_extras={
                "cartao_cnpj_cnpj_confere": cnpj_confere if texto_extraido or recurso.conteudo else None,
                "cartao_cnpj_razao_social_confere": razao_confere if texto_extraido or recurso.conteudo else None,
                "cartao_cnpj_fonte": link_normalizado,
            },
        )

    def _avaliar_canal(
        self,
        inscricao: Inscricao,
        url: str,
        recurso: RecursoRemoto,
    ) -> dict[str, Any]:
        parsed = urlparse(url)
        dominio = (parsed.netloc or "").lower()
        dominio_base = dominio.replace("www.", "")
        tipo_canal = "rede_social" if dominio_base in SOCIAL_DOMAINS else "website"
        tokens = self._tokens_empresa(inscricao.empresa_nome, inscricao.obter_campo("razao_social"))
        slug = normalizar_texto(parsed.path.replace("/", " "))
        texto_referencia = " ".join(
            item
            for item in [
                dominio_base,
                slug,
                recurso.title or "",
                recurso.texto[:4000] if recurso.texto else "",
            ]
            if item
        )

        score = 0.0
        justificativas = [f"URL normalizada para {url}."]
        if dominio_base:
            score += 0.25
        if recurso.erro is None:
            score += 0.25
        else:
            justificativas.append(f"Falha ao acessar o canal: {recurso.erro}.")

        hits = self._contar_tokens_empresa(texto_referencia, tokens)
        if hits >= 2:
            score += 0.35
            justificativas.append("Foram encontrados multiplos sinais de associacao do canal com a empresa.")
        elif hits == 1:
            score += 0.2
            justificativas.append("Foi encontrado ao menos um sinal de associacao do canal com a empresa.")
        else:
            justificativas.append("Nao foram encontrados sinais claros de associacao no dominio, slug ou conteudo.")

        if tipo_canal == "rede_social" and parsed.path.strip("/"):
            score += 0.1
            justificativas.append("Link aponta para um perfil especifico em rede social.")

        valor: bool | None
        revisao_humana = False
        if score >= 0.65:
            valor = True
        elif score >= 0.4:
            valor = True
            revisao_humana = True
        elif recurso.erro:
            valor = None
            revisao_humana = True
        else:
            valor = False

        if self.openai.disponivel and recurso.texto and (revisao_humana or valor is False):
            parecer_ia = self.openai.analisar_canal(
                empresa_nome=inscricao.empresa_nome,
                razao_social=inscricao.obter_campo("razao_social"),
                url=url,
                titulo=recurso.title,
                html_lang=recurso.html_lang,
                texto=recurso.texto[:6000],
            )
            if parecer_ia:
                justificativas.append(f"IA: {parecer_ia.get('explanation', '').strip()}")
                if parecer_ia.get("official_channel") is True:
                    valor = True
                    revisao_humana = revisao_humana or score < 0.75
                elif parecer_ia.get("official_channel") is False and score < 0.5:
                    valor = False
                    revisao_humana = False

        return {
            "valor": valor,
            "revisao_humana": revisao_humana,
            "justificativa": " ".join(justificativas).strip(),
            "tipo_canal": tipo_canal,
        }

    def _registrar_resultado_verificacao(
        self,
        inscricao: Inscricao,
        *,
        prefixo: str,
        valor: Any,
        revisao_humana: bool,
        valor_observado: Any,
        justificativa: str,
        campos_extras: dict[str, Any] | None = None,
    ) -> None:
        inscricao.definir_campo(prefixo, valor)
        inscricao.definir_campo(f"{prefixo}_revisao_humana", revisao_humana)
        inscricao.definir_campo(f"{prefixo}_valor_observado", valor_observado)
        inscricao.definir_campo(f"{prefixo}_justificativa", justificativa)

        for chave, valor_extra in (campos_extras or {}).items():
            inscricao.definir_campo(chave, valor_extra)

    def _obter_recurso(self, url: str) -> RecursoRemoto:
        if url in self._cache_recursos:
            return self._cache_recursos[url]

        if self._atingiu_limite_recursos_remotos(url):
            recurso = RecursoRemoto(
                url=url,
                erro=(
                    "Limite de verificacoes remotas atingido nesta execucao. "
                    "Os demais casos devem seguir para revisao humana."
                ),
            )
            self._cache_recursos[url] = recurso
            return recurso

        recurso = self._baixar_recurso(url)
        if self._url_eh_remota(url):
            self._recursos_remotos_processados += 1
        self._cache_recursos[url] = recurso
        return recurso

    def _normalizar_limite_recursos(self, valor: Any) -> int | None:
        if valor in (None, "", False):
            return None
        try:
            limite = int(valor)
        except (TypeError, ValueError):
            return None
        return limite if limite > 0 else None

    def _atingiu_limite_recursos_remotos(self, url: str) -> bool:
        if not self._url_eh_remota(url):
            return False
        if self.max_recursos_remotos_por_execucao is None:
            return False
        return self._recursos_remotos_processados >= self.max_recursos_remotos_por_execucao

    def _url_eh_remota(self, url: str) -> bool:
        parsed = urlparse(url)
        return parsed.scheme in {"http", "https"}

    def _baixar_recurso(self, url: str) -> RecursoRemoto:
        parsed = urlparse(url)
        if parsed.scheme == "file":
            caminho = Path(parsed.path)
            return self._ler_arquivo_local(url, caminho)

        if parsed.scheme in {"", None} and Path(url).exists():
            return self._ler_arquivo_local(url, Path(url))

        headers = {"User-Agent": self.user_agent, "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.7,es;q=0.6"}
        request = Request(url, headers=headers, method="GET")
        recurso = RecursoRemoto(url=url, final_url=url)

        try:
            context = ssl.create_default_context()
            with urlopen(request, timeout=self.timeout_segundos, context=context) as resposta:
                conteudo = resposta.read(self.max_download_bytes + 1)
                if len(conteudo) > self.max_download_bytes:
                    conteudo = conteudo[: self.max_download_bytes]
                recurso.final_url = resposta.geturl()
                recurso.status_code = getattr(resposta, "status", None)
                recurso.content_type = resposta.headers.get_content_type()
                recurso.conteudo = conteudo
                if self._deve_tratar_como_texto(recurso.content_type, recurso.final_url):
                    texto = self._decodificar_texto(conteudo)
                    recurso.texto = texto
                    if recurso.content_type and "html" in recurso.content_type:
                        extrator = _ExtratorHTML()
                        extrator.feed(texto)
                        recurso.html_lang = extrator.html_lang
                        recurso.title = extrator.title.strip() or None
                        recurso.texto = extrator.texto
        except HTTPError as exc:
            recurso.erro = f"HTTP {exc.code}"
        except URLError as exc:
            recurso.erro = str(exc.reason)
        except Exception as exc:
            recurso.erro = str(exc)

        return recurso

    def _ler_arquivo_local(self, url: str, caminho: Path) -> RecursoRemoto:
        recurso = RecursoRemoto(url=url, final_url=str(caminho))
        try:
            conteudo = caminho.read_bytes()
            recurso.conteudo = conteudo[: self.max_download_bytes]
            recurso.content_type = mimetypes.guess_type(caminho.name)[0] or "application/octet-stream"
            if self._deve_tratar_como_texto(recurso.content_type, caminho.name):
                texto = self._decodificar_texto(recurso.conteudo)
                recurso.texto = texto
                if recurso.content_type and "html" in recurso.content_type:
                    extrator = _ExtratorHTML()
                    extrator.feed(texto)
                    recurso.html_lang = extrator.html_lang
                    recurso.title = extrator.title.strip() or None
                    recurso.texto = extrator.texto
        except Exception as exc:
            recurso.erro = str(exc)
        return recurso

    def _extrair_texto_documento(self, recurso: RecursoRemoto) -> str | None:
        if recurso.texto:
            return recurso.texto
        if recurso.content_type and "pdf" in recurso.content_type and recurso.conteudo:
            try:
                reader = PdfReader(BytesIO(recurso.conteudo))
                textos = [pagina.extract_text() or "" for pagina in reader.pages]
                texto = " ".join(textos).strip()
                return texto or None
            except Exception:
                return None
        return None

    def _detectar_idiomas(self, recurso: RecursoRemoto) -> list[str]:
        idiomas: list[str] = []
        if recurso.html_lang:
            html_lang = normalizar_texto(recurso.html_lang)
            if html_lang.startswith("en"):
                idiomas.append("en")
            elif html_lang.startswith("es"):
                idiomas.append("es")
            elif html_lang.startswith("pt"):
                idiomas.append("pt")

        texto = normalizar_texto(recurso.texto or "")
        if not texto:
            return idiomas

        palavras = set(re.findall(r"[a-z]{2,}", texto))
        score_por_idioma = {
            idioma: len(palavras.intersection(stopwords))
            for idioma, stopwords in LANGUAGE_STOPWORDS.items()
        }
        for idioma, score in score_por_idioma.items():
            if score >= 2 and idioma not in idiomas:
                idiomas.append(idioma)
        return idiomas

    def _validar_cnpj(self, cnpj: str) -> bool:
        if len(cnpj) != 14 or len(set(cnpj)) == 1:
            return False

        numeros = [int(digito) for digito in cnpj]
        pesos_1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        pesos_2 = [6] + pesos_1

        digito_1 = self._calcular_digito_cnpj(numeros[:12], pesos_1)
        digito_2 = self._calcular_digito_cnpj(numeros[:13], pesos_2)
        return numeros[12] == digito_1 and numeros[13] == digito_2

    def _calcular_digito_cnpj(self, numeros: list[int], pesos: list[int]) -> int:
        soma = sum(numero * peso for numero, peso in zip(numeros, pesos))
        resto = soma % 11
        return 0 if resto < 2 else 11 - resto

    def _normalizar_url_publica(self, valor: Any) -> str | None:
        if not self._tem_valor(valor):
            return None

        texto = str(valor).strip()
        texto_normalizado = normalizar_texto(texto)
        if texto_normalizado in PLACEHOLDER_VALUES:
            return None

        if texto.startswith("www."):
            return f"https://{texto}"
        if re.match(r"^https?://", texto, flags=re.IGNORECASE):
            return texto
        if re.match(r"^[\w.-]+\.[a-z]{2,}(/.*)?$", texto, flags=re.IGNORECASE):
            return f"https://{texto}"
        return None

    def _normalizar_link_documento(self, valor: Any) -> str | None:
        url = self._normalizar_url_publica(valor)
        if url is None:
            return None

        parsed = urlparse(url)
        dominio = (parsed.netloc or "").lower()
        if "drive.google.com" in dominio:
            correspondencia = re.search(r"/file/d/([^/]+)/", parsed.path)
            if correspondencia:
                arquivo_id = correspondencia.group(1)
                return f"https://drive.google.com/uc?export=download&id={quote_plus(arquivo_id)}"
        return url

    def _deve_tratar_como_texto(self, content_type: str | None, url: str | None) -> bool:
        if content_type and (
            content_type.startswith("text/")
            or "json" in content_type
            or "xml" in content_type
            or "html" in content_type
        ):
            return True

        extensao = Path(urlparse(url or "").path).suffix.lower()
        return extensao in {".html", ".htm", ".txt", ".xml", ".json"}

    def _decodificar_texto(self, conteudo: bytes) -> str:
        for encoding in ("utf-8", "latin-1", "cp1252"):
            try:
                return conteudo.decode(encoding)
            except UnicodeDecodeError:
                continue
        return conteudo.decode("utf-8", errors="ignore")

    def _tokens_empresa(self, *nomes: Any) -> list[str]:
        tokens: list[str] = []
        for nome in nomes:
            texto = normalizar_texto(nome)
            for token in re.findall(r"[a-z0-9]{3,}", texto):
                if token in LEGAL_ENTITY_TOKENS:
                    continue
                if token not in tokens:
                    tokens.append(token)
        return tokens

    def _contar_tokens_empresa(self, texto: str, tokens: list[str]) -> int:
        texto_normalizado = normalizar_texto(texto)
        return sum(1 for token in tokens if token and token in texto_normalizado)

    def _texto_corresponde_empresa(self, texto: str, empresa_nome: str, razao_social: str | None) -> bool:
        tokens = self._tokens_empresa(empresa_nome, razao_social)
        if not tokens:
            return False
        hits = self._contar_tokens_empresa(texto, tokens)
        minimo = 1 if len(tokens) <= 2 else 2
        return hits >= minimo

    def _apenas_digitos(self, valor: Any) -> str:
        if valor is None:
            return ""
        return re.sub(r"\D+", "", str(valor))

    def _formatar_cnpj(self, cnpj: str) -> str:
        if len(cnpj) != 14:
            return cnpj
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

    def _tem_valor(self, valor: Any) -> bool:
        return valor is not None and str(valor).strip() != ""

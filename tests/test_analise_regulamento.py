from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from services.analise_regulamento import executar_analise_regulamento


class AnaliseRegulamentoTestCase(unittest.TestCase):
    def test_deve_gerar_excel_de_analise_a_partir_de_txt_e_planilha(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            base = Path(diretorio_temporario)
            regulamento = base / "regulamento.txt"
            planilha = base / "inscricoes.csv"
            saida = base / "analise.xlsx"

            regulamento.write_text(
                (
                    "O regulamento exige website em ingles ou espanhol, cartao CNPJ, "
                    "atendimento PEIEX, consultoria Sebrae e indicacao de NCM."
                ),
                encoding="utf-8",
            )
            planilha.write_text(
                "\n".join(
                    [
                        "Inscrição (Código),Nome Fantasia,CNPJ,Status Financeiro,Informe o website da empresa ou sua conta nas redes sociais: (C016)",
                        "INS-1,Empresa A,12.345.678/0001-99,Adimplente,https://empresa-a.com",
                    ]
                ),
                encoding="utf-8",
            )

            resumo = executar_analise_regulamento(planilha, regulamento, saida)

            self.assertEqual(resumo.modo, "analise_regulamento")
            self.assertTrue(saida.exists())
            self.assertGreaterEqual(resumo.total_criterios_detectados, 4)
            self.assertGreaterEqual(resumo.total_mapeamentos_sugeridos, 4)

            workbook = load_workbook(saida, read_only=True, data_only=True)
            self.assertIn("criterios_detectados", workbook.sheetnames)
            self.assertIn("config_sugerida", workbook.sheetnames)
            workbook.close()


if __name__ == "__main__":
    unittest.main()

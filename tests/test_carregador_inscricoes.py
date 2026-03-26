from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from services.carregador_inscricoes import carregar_inscricoes


class CarregadorInscricoesTestCase(unittest.TestCase):
    def test_deve_carregar_xlsx_com_mapeamento_de_campos_e_aba(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            caminho = Path(diretorio_temporario) / "entrada.xlsx"
            workbook = Workbook()
            worksheet_padrao = workbook.active
            worksheet_padrao.title = "Resumo"
            worksheet_padrao["A1"] = "ignorar"

            worksheet = workbook.create_sheet("Ranqueamento")
            worksheet.append(
                [
                    "CNPJ",
                    "Nome Fantasia",
                    "Data de Inscricao",
                    "Pontuacao PEIEX",
                ]
            )
            worksheet.append(
                [
                    "12.345.678/0001-99",
                    "Empresa Teste",
                    "2026-01-20 10:00:00",
                    1,
                ]
            )
            workbook.save(caminho)
            workbook.close()

            inscricoes = carregar_inscricoes(
                caminho,
                sheet_name="Ranqueamento",
                field_map={
                    "inscricao_id": ["CNPJ"],
                    "empresa_nome": ["Nome Fantasia"],
                    "data_submissao": ["Data de Inscricao"],
                    "pontuacao_peiex_legado": ["Pontuacao PEIEX"],
                },
            )

            self.assertEqual(len(inscricoes), 1)
            inscricao = inscricoes[0]
            self.assertEqual(inscricao.inscricao_id, "12.345.678/0001-99")
            self.assertEqual(inscricao.empresa_nome, "Empresa Teste")
            self.assertEqual(inscricao.obter_campo("pontuacao_peiex_legado"), 1)
            self.assertIsNotNone(inscricao.data_submissao)

    def test_deve_cair_para_aba_ativa_quando_nome_configurado_nao_existe(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            caminho = Path(diretorio_temporario) / "entrada.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "jornada-exportadora-autopecas-"
            worksheet.append(["Inscrição (Código)", "Nome Fantasia", "CNPJ"])
            worksheet.append(["INS-1", "Empresa CRM", "12.345.678/0001-99"])
            workbook.save(caminho)
            workbook.close()

            inscricoes = carregar_inscricoes(
                caminho,
                sheet_name="Ranqueamento",
                field_map={
                    "inscricao_id": ["Inscrição (Código)"],
                    "empresa_nome": ["Nome Fantasia"],
                    "cnpj": ["CNPJ"],
                },
            )

            self.assertEqual(len(inscricoes), 1)
            self.assertEqual(inscricoes[0].inscricao_id, "INS-1")
            self.assertEqual(inscricoes[0].empresa_nome, "Empresa CRM")

    def test_deve_mapear_campo_por_alias_parcial_quando_header_tem_codigo_do_formulario(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            caminho = Path(diretorio_temporario) / "entrada.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Ranqueamento"
            worksheet.append(
                [
                    "Inscrição (Código)",
                    "Nome Fantasia",
                    "Informe o website da empresa ou sua conta nas redes sociais: (C016)",
                ]
            )
            worksheet.append(["INS-1", "Empresa CRM", "https://empresa.example.com"])
            workbook.save(caminho)
            workbook.close()

            inscricoes = carregar_inscricoes(
                caminho,
                sheet_name="Ranqueamento",
                field_map={
                    "inscricao_id": ["Inscrição (Código)"],
                    "empresa_nome": ["Nome Fantasia"],
                    "website_empresa": ["Informe o website da empresa ou sua conta nas redes sociais"],
                },
            )

            self.assertEqual(len(inscricoes), 1)
            self.assertEqual(
                inscricoes[0].obter_campo("website_empresa"),
                "https://empresa.example.com",
            )

    def test_deve_rejeitar_workbook_que_ja_e_saida_do_sistema(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            caminho = Path(diretorio_temporario) / "saida_sistema.xlsx"
            workbook = Workbook()
            workbook.active.title = "ranking_final"
            workbook.create_sheet("avaliacao_por_criterio")
            workbook.create_sheet("pendencias_revisao")
            workbook.create_sheet("inscricoes_brutas")
            workbook.save(caminho)
            workbook.close()

            with self.assertRaisesRegex(ValueError, "saida do proprio sistema"):
                carregar_inscricoes(caminho)


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from models import Inscricao
from services.exportador_excel import ExportadorExcel
from services.motor_regras import MotorRegras
from services.ranking_service import RankingService


REGULAMENTO_TESTE = {
    "elegibilidade": [
        {
            "id": "aceite_regulamento",
            "tipo": "equals",
            "campo": "aceite_regulamento",
            "valor_esperado": True,
        }
    ],
    "pontuacao": {
        "nota_minima_classificacao": 10,
        "criterios": [
            {
                "id": "experiencia_exportadora",
                "tipo": "range_score",
                "campo": "anos_experiencia_exportacao",
                "faixas": [{"min": 0, "score": 10}],
            },
            {
                "id": "maturidade_exportadora",
                "tipo": "value_map",
                "campo": "maturidade_exportadora",
                "mapa_valores": {"alta": 5},
                "revisao_humana": True,
                "contabilizar_sugestao_na_nota": True,
            }
        ],
    },
    "desempate": [],
}


class ExportadorExcelTestCase(unittest.TestCase):
    def test_deve_gerar_workbook_com_abas_esperadas(self) -> None:
        motor = MotorRegras(REGULAMENTO_TESTE)
        ranking_service = RankingService(REGULAMENTO_TESTE)
        exportador = ExportadorExcel()

        inscricao = Inscricao.from_row(
            {
                "inscricao_id": "1",
                "empresa_nome": "Empresa A",
                "razao_social": "Empresa A Ltda",
                "cnpj": "12.345.678/0001-99",
                "aceite_regulamento": "Sim",
                "anos_experiencia_exportacao": "3",
                "maturidade_exportadora": "Alta",
            }
        )

        ranking = ranking_service.classificar([motor.avaliar(inscricao)])

        with tempfile.TemporaryDirectory() as diretorio_temporario:
            saida = Path(diretorio_temporario) / "ranking.xlsx"
            exportador.exportar(saida, [inscricao], ranking)

            workbook = load_workbook(saida)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "ranking_final",
                    "avaliacao_por_criterio",
                    "pendencias_revisao",
                    "inscricoes_brutas",
                ],
            )
            self.assertEqual(workbook["ranking_final"]["A2"].value, 1)
            self.assertEqual(workbook["ranking_final"]["B2"].value, "Empresa A")
            self.assertEqual(workbook["ranking_final"]["C2"].value, "Empresa A Ltda")
            self.assertEqual(workbook["ranking_final"]["D2"].value, "12.345.678/0001-99")
            self.assertEqual(workbook["ranking_final"]["H2"].value, "classificada_com_revisao_pendente")
            self.assertEqual(workbook["pendencias_revisao"]["A2"].value, "1")
            self.assertEqual(workbook["ranking_final"].freeze_panes, "A2")
            self.assertEqual(workbook["ranking_final"].auto_filter.ref, "A1:K2")
            self.assertEqual(workbook["ranking_final"].sheet_properties.tabColor.rgb, "002FD126")
            self.assertEqual(workbook["avaliacao_por_criterio"].sheet_state, "visible")
            self.assertEqual(workbook["pendencias_revisao"].sheet_state, "visible")
            self.assertEqual(workbook["inscricoes_brutas"].sheet_state, "hidden")
            self.assertTrue(workbook["ranking_final"]["A1"].font.bold)
            self.assertEqual(workbook["ranking_final"]["A1"].fill.fgColor.rgb, "00132B8F")
            self.assertEqual(workbook["ranking_final"]["H2"].fill.fgColor.rgb, "00FFF3C8")
            self.assertEqual(workbook["ranking_final"]["E1"].value, "Pontos - Experiencia exportadora")
            self.assertEqual(workbook["ranking_final"]["E2"].value, 10)
            self.assertEqual(workbook["ranking_final"]["E2"].number_format, "0.00")
            self.assertEqual(workbook["ranking_final"]["F1"].value, "Pontos - Maturidade exportadora")
            self.assertEqual(workbook["ranking_final"]["F2"].value, 5)
            self.assertEqual(workbook["ranking_final"]["G1"].value, "Nota total")
            self.assertEqual(workbook["ranking_final"]["G2"].value, 15)
            self.assertEqual(workbook["ranking_final"]["G2"].number_format, "0.00")
            self.assertEqual(workbook["ranking_final"]["J1"].value, "Resumo da classificacao")
            self.assertIn("experiencia_exportadora", workbook["ranking_final"]["J2"].value)
            self.assertTrue(workbook["ranking_final"]["J2"].alignment.wrap_text)
            self.assertEqual(workbook["ranking_final"]["K1"].value, "Revisao humana - Maturidade exportadora")
            self.assertIn("revisao humana", workbook["ranking_final"]["K2"].value.lower())
            self.assertTrue(workbook["ranking_final"]["K2"].alignment.wrap_text)
            self.assertEqual(workbook["ranking_final"]["A2"].fill.fgColor.rgb, "001F8F43")
            self.assertEqual(workbook["ranking_final"]["B2"].fill.fgColor.rgb, "00E3F6E7")
            workbook.close()


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from services.pipeline import executar_ranqueamento


class PipelineTestCase(unittest.TestCase):
    def test_deve_executar_pipeline_e_gerar_excel(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            base = Path(diretorio_temporario)
            entrada = base / "entrada.csv"
            saida = base / "resultado.xlsx"

            entrada.write_text(
                "\n".join(
                    [
                        "inscricao_id,empresa_nome,aceite_regulamento,situacao_cadastral,anos_experiencia_exportacao,quantidade_mercados_prioritarios,maturidade_exportadora",
                        "1,Empresa A,Sim,Ativa,6,3,Alta",
                        "2,Empresa B,Nao,Ativa,1,0,Baixa",
                    ]
                ),
                encoding="utf-8",
            )

            resumo = executar_ranqueamento(
                entrada,
                "config/regulamento_exemplo.json",
                saida,
            )

            self.assertEqual(resumo.total_inscricoes, 2)
            self.assertTrue(saida.exists())
            self.assertIn("classificada_com_revisao_pendente", resumo.contagem_status)

    def test_deve_filtrar_por_status_do_dynamics_excluir_apex_e_gerar_planilha_unica(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            base = Path(diretorio_temporario)
            entrada = base / "entrada.csv"
            saida = base / "resultado.xlsx"
            regulamento = base / "regulamento.json"

            entrada.write_text(
                "\n".join(
                    [
                        "inscricao_id,empresa_nome,razao_social,razao_status,aceite_regulamento,website",
                        "1,Empresa A,Empresa A Ltda,Pendente Avaliação,Sim,https://empresa-a.com",
                        "2,Empresa B,Empresa B Ltda,Reprovada,Sim,https://empresa-b.com",
                        "3,APEX-BRASIL,AGENCIA DE PROMOCAO DE EXPORTACOES DO BRASIL - APEX-BRASIL,Aprovada,Sim,https://apex.example.com",
                        "4,Empresa C,Empresa C Ltda,Pré Aprovada,Sim,https://empresa-c.com",
                    ]
                ),
                encoding="utf-8",
            )

            regulamento.write_text(
                """
                {
                  "nome_programa": "Teste Dynamics",
                  "versao": "1.0",
                  "origem_dados": {
                    "filtros_entrada": [
                      {
                        "id": "status_dynamics_analise",
                        "modo": "include",
                        "tipo": "texto_em_lista",
                        "campo": "razao_status",
                        "valores": ["Aprovada", "Pré Aprovada", "Pre Aprovada", "Pendente Avaliação", "Pendente Avaliacao"]
                      },
                      {
                        "id": "excluir_testes_apexbrasil",
                        "modo": "exclude",
                        "tipo": "texto_contem_algum",
                        "campos": ["empresa_nome", "razao_social"],
                        "valores": ["ApexBrasil", "Apex Brasil"]
                      }
                    ]
                  },
                  "saida_excel": {
                    "somente_ranking_final": true
                  },
                  "elegibilidade": [
                    {
                      "id": "aceite_regulamento",
                      "tipo": "equals",
                      "campo": "aceite_regulamento",
                      "valor_esperado": true
                    }
                  ],
                  "pontuacao": {
                    "nota_minima_classificacao": 1,
                    "criterios": [
                      {
                        "id": "website",
                        "tipo": "presence_score",
                        "campo": "website",
                        "pontuacao": 1
                      }
                    ]
                  },
                  "desempate": []
                }
                """.strip(),
                encoding="utf-8",
            )

            resumo = executar_ranqueamento(entrada, regulamento, saida)

            self.assertEqual(resumo.total_inscricoes, 2)
            workbook = load_workbook(saida, read_only=True, data_only=True)
            self.assertEqual(workbook.sheetnames, ["ranking_final"])
            ranking = workbook["ranking_final"]
            empresas = [ranking["B2"].value, ranking["B3"].value]
            self.assertEqual(empresas, ["Empresa A", "Empresa C"])
            workbook.close()

    def test_deve_deduplicar_por_cnpj_mantendo_primeira_data_de_conclusao(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            base = Path(diretorio_temporario)
            entrada = base / "entrada.csv"
            saida = base / "resultado.xlsx"
            regulamento = base / "regulamento.json"

            entrada.write_text(
                "\n".join(
                    [
                        "inscricao_id,empresa_nome,cnpj,data_submissao,aceite_regulamento,website",
                        "1,Empresa A,12.345.678/0001-95,2026-03-10 10:00:00,Sim,https://empresa-a.com",
                        "2,Empresa A,12.345.678/0001-95,2026-03-09 09:00:00,Sim,https://empresa-a.com",
                        "3,Empresa B,98.765.432/0001-10,2026-03-11 12:00:00,Sim,https://empresa-b.com",
                    ]
                ),
                encoding="utf-8",
            )

            regulamento.write_text(
                """
                {
                  "nome_programa": "Teste Deduplicacao",
                  "versao": "1.0",
                  "origem_dados": {
                    "deduplicacao": {
                      "tipo": "cnpj_primeira_submissao",
                      "campo": "cnpj",
                      "campo_data": "data_submissao"
                    }
                  },
                  "saida_excel": {
                    "somente_ranking_final": true
                  },
                  "elegibilidade": [
                    {
                      "id": "aceite_regulamento",
                      "tipo": "equals",
                      "campo": "aceite_regulamento",
                      "valor_esperado": true
                    }
                  ],
                  "pontuacao": {
                    "nota_minima_classificacao": 1,
                    "criterios": [
                      {
                        "id": "website",
                        "tipo": "presence_score",
                        "campo": "website",
                        "pontuacao": 1
                      }
                    ]
                  },
                  "desempate": []
                }
                """.strip(),
                encoding="utf-8",
            )

            resumo = executar_ranqueamento(entrada, regulamento, saida)

            self.assertEqual(resumo.total_inscricoes, 2)
            workbook = load_workbook(saida, read_only=True, data_only=True)
            ranking = workbook["ranking_final"]
            empresas = [ranking["B2"].value, ranking["B3"].value]
            self.assertEqual(empresas, ["Empresa A", "Empresa B"])
            workbook.close()


if __name__ == "__main__":
    unittest.main()

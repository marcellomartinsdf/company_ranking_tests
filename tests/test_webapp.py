from __future__ import annotations

import io
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from webapp import create_app


class WebAppTestCase(unittest.TestCase):
    def test_deve_processar_upload_e_exibir_link_download(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            base = Path(diretorio_temporario)
            config_dir = base / "config"
            storage_dir = base / "storage"
            config_dir.mkdir(parents=True, exist_ok=True)
            storage_dir.mkdir(parents=True, exist_ok=True)

            regulamento = {
                "nome_programa": "Teste Web",
                "versao": "1.0",
                "elegibilidade": [
                    {
                        "id": "aceite_regulamento",
                        "tipo": "equals",
                        "campo": "aceite_regulamento",
                        "valor_esperado": True,
                    }
                ],
                "pontuacao": {
                    "nota_minima_classificacao": 1,
                    "criterios": [
                        {
                            "id": "website",
                            "tipo": "presence_score",
                            "campo": "website",
                            "pontuacao": 1,
                        }
                    ],
                },
                "desempate": [],
            }
            (config_dir / "teste.json").write_text(
                json.dumps(regulamento),
                encoding="utf-8",
            )

            app = create_app()
            app.config["TESTING"] = True
            app.config["CONFIG_DIR"] = str(config_dir)
            app.config["STORAGE_DIR"] = str(storage_dir)
            app.config["JOB_STORAGE_BACKEND"] = "local"
            client = app.test_client()

            resposta = client.post(
                "/processar",
                data={
                    "planilha": (
                        io.BytesIO(
                            (
                                "inscricao_id,empresa_nome,aceite_regulamento,website\n"
                                "1,Empresa A,Sim,https://empresa-a.com\n"
                            ).encode("utf-8")
                        ),
                        "inscricoes.csv",
                    ),
                },
                content_type="multipart/form-data",
            )

            self.assertEqual(resposta.status_code, 200)
            self.assertIn("Baixar Excel", resposta.get_data(as_text=True))
            arquivos = list(storage_dir.glob("**/*.xlsx"))
            self.assertTrue(arquivos)

    def test_deve_usar_criterios_do_regulamento_pre_selecionado_na_saida(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            base = Path(diretorio_temporario)
            config_dir = base / "config"
            storage_dir = base / "storage"
            config_dir.mkdir(parents=True, exist_ok=True)
            storage_dir.mkdir(parents=True, exist_ok=True)

            regulamento_padrao = {
                "nome_programa": "Jornada Exportadora Autopartes - Argentina e Peru 2026",
                "versao": "1.0",
                "elegibilidade": [],
                "pontuacao": {
                    "nota_minima_classificacao": 0,
                    "criterios": [
                        {
                            "id": "website",
                            "nome": "Website institucional",
                            "tipo": "presence_score",
                            "campo": "website",
                            "pontuacao": 1,
                        }
                    ],
                },
                "desempate": [],
            }
            regulamento_construcao = {
                "nome_programa": "Exporta Mais Brasil - Construcao 2026",
                "versao": "1.0",
                "elegibilidade": [],
                "pontuacao": {
                    "nota_minima_classificacao": 0,
                    "criterios": [
                        {
                            "id": "exportacao_direta",
                            "nome": "Exportacao direta nos ultimos 2 anos",
                            "tipo": "binary_score",
                            "campo": "exportou_diretamente_ultimos_2_anos",
                            "valor_esperado": True,
                            "pontuacao": 2,
                        },
                        {
                            "id": "lideranca_feminina",
                            "nome": "Lideranca feminina",
                            "tipo": "binary_score",
                            "campo": "liderada_por_mulher",
                            "valor_esperado": True,
                            "pontuacao": 1,
                        },
                    ],
                },
                "desempate": [],
            }
            (config_dir / "regulamento_jornada_exportadora_autopartes_2026.json").write_text(
                json.dumps(regulamento_padrao),
                encoding="utf-8",
            )
            (config_dir / "regulamento_exporta_mais_brasil_construcao_2026.json").write_text(
                json.dumps(regulamento_construcao),
                encoding="utf-8",
            )

            app = create_app()
            app.config["TESTING"] = True
            app.config["CONFIG_DIR"] = str(config_dir)
            app.config["STORAGE_DIR"] = str(storage_dir)
            app.config["JOB_STORAGE_BACKEND"] = "local"
            client = app.test_client()

            resposta = client.post(
                "/processar",
                data={
                    "regulamento_preset": "regulamento_exporta_mais_brasil_construcao_2026.json",
                    "planilha": (
                        io.BytesIO(
                            (
                                "inscricao_id,empresa_nome,exportou_diretamente_ultimos_2_anos,liderada_por_mulher\n"
                                "1,Empresa A,Sim,Sim\n"
                            ).encode("utf-8")
                        ),
                        "inscricoes.csv",
                    ),
                },
                content_type="multipart/form-data",
            )

            self.assertEqual(resposta.status_code, 200)
            conteudo = resposta.get_data(as_text=True)
            self.assertIn("Exporta Mais Brasil - Construcao 2026", conteudo)

            arquivos = list(storage_dir.glob("**/*.xlsx"))
            self.assertTrue(arquivos)

            workbook = load_workbook(arquivos[0], read_only=True, data_only=True)
            ranking_final = workbook["ranking_final"]
            cabecalho = [celula.value for celula in ranking_final[1]]
            self.assertIn("Pontos - Exportacao direta nos ultimos 2 anos", cabecalho)
            self.assertIn("Pontos - Lideranca feminina", cabecalho)
            self.assertNotIn("Pontos - Website institucional", cabecalho)
            workbook.close()

    def test_deve_processar_planilha_com_aba_diferente_no_upload_web(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            base = Path(diretorio_temporario)
            config_dir = base / "config"
            storage_dir = base / "storage"
            config_dir.mkdir(parents=True, exist_ok=True)
            storage_dir.mkdir(parents=True, exist_ok=True)

            regulamento = {
                "nome_programa": "Teste Web Dynamics",
                "versao": "1.0",
                "origem_dados": {
                    "sheet_name": "Ranqueamento",
                    "mapeamento_campos": {
                        "inscricao_id": ["Inscrição (Código)"],
                        "empresa_nome": ["Nome Fantasia"],
                        "cnpj": ["CNPJ"],
                    },
                },
                "elegibilidade": [],
                "pontuacao": {"nota_minima_classificacao": 0, "criterios": []},
                "desempate": [],
            }
            (config_dir / "teste_dynamics.json").write_text(
                json.dumps(regulamento),
                encoding="utf-8",
            )

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "jornada-exportadora-autopecas-"
            worksheet.append(["Inscrição (Código)", "Nome Fantasia", "CNPJ"])
            worksheet.append(["INS-1", "Empresa CRM", "12.345.678/0001-99"])
            buffer = io.BytesIO()
            workbook.save(buffer)
            workbook.close()
            buffer.seek(0)

            app = create_app()
            app.config["TESTING"] = True
            app.config["CONFIG_DIR"] = str(config_dir)
            app.config["STORAGE_DIR"] = str(storage_dir)
            app.config["JOB_STORAGE_BACKEND"] = "local"
            client = app.test_client()

            resposta = client.post(
                "/processar",
                data={
                    "planilha": (buffer, "inscricoes_dynamics.xlsx"),
                },
                content_type="multipart/form-data",
            )

            self.assertEqual(resposta.status_code, 200)
            self.assertIn("Baixar Excel", resposta.get_data(as_text=True))

    def test_deve_mostrar_regulamento_padrao_na_tela_inicial(self) -> None:
        with tempfile.TemporaryDirectory() as diretorio_temporario:
            base = Path(diretorio_temporario)
            config_dir = base / "config"
            config_dir.mkdir(parents=True, exist_ok=True)

            regulamento = {
                "nome_programa": "Jornada Exportadora Autopartes - Argentina e Peru 2026",
                "versao": "3.1",
                "elegibilidade": [],
                "pontuacao": {"nota_minima_classificacao": 0, "criterios": []},
                "desempate": [],
            }
            (config_dir / "regulamento_jornada_exportadora_autopartes_2026.json").write_text(
                json.dumps(regulamento),
                encoding="utf-8",
            )

            app = create_app()
            app.config["TESTING"] = True
            app.config["CONFIG_DIR"] = str(config_dir)
            client = app.test_client()

            resposta = client.get("/")
            self.assertEqual(resposta.status_code, 200)
            conteudo = resposta.get_data(as_text=True)
            self.assertIn("Jornada Exportadora Autopartes - Argentina e Peru 2026", conteudo)
            self.assertIn("regulamento_preset", conteudo)
            self.assertNotIn("regulamento_customizado", conteudo)
            self.assertNotIn("regulamento_documento", conteudo)


if __name__ == "__main__":
    unittest.main()
